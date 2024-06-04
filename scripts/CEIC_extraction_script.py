#!/usr/bin/env python
# coding: utf-8

# In[1]:


import warnings
warnings.filterwarnings('ignore')
import PySimpleGUI as sg
sg.theme('DarkBlue13')
from ceic_api_client.pyceic import Ceic
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from tqdm import tqdm
import re
import subprocess
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timezone
import xlwings as xw
import shutil
import math
import statistics
from functools import reduce  
import os, sys
import cloudscraper
import PIL
import io
from pathlib import Path


# In[2]:


#Extract series dataframe from CEIC using the series ID

def get_ceic_data(series_id, start_date=''):
    try:
        if start_date == '':
            series_result = Ceic.series(series_id)
        else:
            series_result = Ceic.series(series_id, start_date=start_date)
        series = series_result.data[0]
        data = {'date': [], 'value': []}
        for i in series.time_points:
            data['date'].append(i._date)
            data['value'].append(i.value)
        df = pd.DataFrame(data)
        final_data = ({'Name': series.metadata.name, 'Region': series.metadata.country.name, 
                          'Frequency': series.metadata.frequency.name, 'Unit': series.metadata.unit.name, 
                          'Source': series.metadata.source.name, 'Series ID': series.entity_id, 
                          'First Obs. Date': series.metadata.start_date, 'Last Obs. Date': series.metadata.end_date,
                          'Last Update Time': series.metadata.last_update_time, 'Data': df})
        return final_data
    except:
        #print('No dataset found for series', str(series_id))
        return np.nan


# In[3]:


def write_excel(filename, sheet_name, df):
    if not os.path.isfile(filename):
        wb = Workbook()
        ws = wb.active
        wb.save(filename)
        wb.close()
    og_cols = list(df.columns)
    cols=pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique(): 
        cols[cols[cols == dup].index.values.tolist()] = [dup + '.' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
    df.columns=cols
    wb = load_workbook(filename)
    if sheet_name in wb.get_sheet_names():
        sheet_to_del=wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet_to_del)
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    w_exc_dict = {}
    cols = list(df.columns)
    for c in cols:
        w_exc_dict[c] = list(df[c])
    col_count = 1
    for d in w_exc_dict.keys():
        ws.cell(row=1, column=col_count).value = og_cols[col_count-1]
        for i in range(len(w_exc_dict[d])):
            try:
                w_exc_dict[d][i] = w_exc_dict[d][i].replace(tzinfo=None)
            except:
                pass
            ws.cell(row=i+2, column=col_count).value = w_exc_dict[d][i]
        try:
            m_row = ws.max_row
            for i in range(m_row+1):
                try:
                    ws.cell(row=i+1, column=col_count).value = ws.cell(row=i+1, column=col_count).value.strftime("%m/%d/%Y")
                except:
                    pass
        except:
            pass
        col_count+=1
    wb.save(filename)
    wb.close()


# In[4]:


#Resize an image from the web based on the resize parameter.
def resize_img(url, resize):
    data_url = (cloudscraper.create_scraper(browser={"browser": "firefox", "platform": "windows", "mobile": False})
        .get(url)
        .content)
    img = PIL.Image.open(io.BytesIO(data_url))
    cur_width, cur_height = img.size
    new_width, new_height = resize
    scale = min(new_height / cur_height, new_width / cur_width)
    img = img.resize((int(cur_width * scale), int(cur_height * scale)), PIL.Image.ANTIALIAS)
    png_bio = io.BytesIO()
    img.save(png_bio, format="PNG")
    png_data = png_bio.getvalue()
    return png_data

#Open the home page screen of the GUI.
def UI_home_page_design():
    png_data_adb = resize_img('https://upload.wikimedia.org/wikipedia/commons/thumb/4/43/Asian_Development_Bank_logo.svg/2048px-Asian_Development_Bank_logo.svg.png', (118,118))
    png_data_banner = resize_img('https://asianbondsonline.adb.org/macroeconomictracker/images/woman-planting-with-charts-trackingasia2.jpg', (300, 300))

    intro_txt = "This tool allows you to extract data from CEIC using the provided template to quickly prepare your data for EAI computations."
    welcome_txt = sg.Text('CEIC Extraction Interface', font=('Calibri Bold', 14, "bold"), expand_x=True, justification='center')
    intro_txt = sg.Text(intro_txt, size=(40, None), font=('Calibri', 12), expand_x=True, justification='center')
    continue_button = sg.Button('Continue', key='CONTINUE_HOME', font=('Calibri', 8), expand_x=True)
    exit = sg.Button('Exit', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True)
    adb_banner = sg.Image(data=png_data_adb, key='-ADB-', pad=(0,0))
    banner = sg.Image(data=png_data_banner, key='-IMAGE-', pad=(0,0))
    layout = [
        [adb_banner, banner],
        [welcome_txt],
        [intro_txt],
        [continue_button, exit]
    ]
    window = sg.Window('CEIC Extraction', layout, grab_anywhere=True, resizable=True, font=('Calibri',12))
    while True:
        event, values = window.read()
        if event == 'CONTINUE_HOME':
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break
    window.close()
    return event


# In[5]:


def UI_ceic_login():
    
    default_user = ''
    default_pass = ''
    
    back_button = sg.Button('Back', key='BACK_CHECK_CEIC', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Enter CEIC Login Credentials', font=('Calibri Bold', 12), expand_x=True, justification='center')
    ceicuser_txt = sg.Text('Username:', key='-OUT-', font=('Calibri', 12), expand_x=True, justification='left')
    ceicpass_txt = sg.Text('Password:', key='-OUT-', font=('Calibri', 12), expand_x=True, justification='left')
    ceicuser = sg.Input(default_user, enable_events=True, key='CEICUSER', font=('Calibri', 8), justification='center')
    ceicpass = sg.Input(default_pass, enable_events=True, key='CEICPASS', font=('Calibri', 8), justification='center', password_char = '*')
    ok_txt = sg.Button('Ok', key='OK_LOGIN', font=('Calibri', 8), expand_x=True)
    exit = sg.Button('Exit', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True)
    
    layout = [
        [back_button],
        [title_txt],
        [ceicuser_txt, ceicuser],
        [ceicpass_txt, ceicpass],
        [ok_txt, exit]
    ]
    window = sg.Window('Economic Activity Index: CEIC Login', layout, grab_anywhere=True,resizable=True)
    while True:
        event, values = window.read()
        if event == 'OK_LOGIN':
            try:
                ceic_username = values['CEICUSER']
                ceic_password = values['CEICPASS']
                Ceic.login(ceic_username, ceic_password)
                break
            except Exception as e:
                #error_msg = e.message
                sg.Popup('CEIC login failed.\nError Message: '+str(e), font=('Calibri', 8), grab_anywhere=True, button_justification='center')
        elif event == 'BACK_CHECK_CEIC':
            window.close()
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            window.close()
            break
    window.close()
    return event


# In[6]:


#Open the screen to browse the local device and select the excel file to be used in the analysis.
#Also checks if the file is formatted correctly
def UI_add_excel_data(): 
    layout = [
                  [sg.Button("Back", key='BACK_HOME', expand_x=False, font=('Calibri',8))],
                  [sg.Text(f'Locate files and folders for CEIC data extraction.',
                          justification='center', expand_x=True, font=('Calibri Bold', 12))],
                  [sg.Text('CEIC Template File', font=('Calibri Bold',12)), sg.Input(key='Input_1', expand_x=True), 
                   sg.FileBrowse(font=('Calibri', 8), tooltip='Click to locate the Excel File template.')],
                  [sg.Text('Output Folder', font=('Calibri Bold',12)), sg.Input(key='Input_2', expand_x=True), 
                   sg.FolderBrowse(font=('Calibri', 8), tooltip='Click to choose location of final outputs.')],
                  [sg.Text('Output File Name', font=('Calibri Bold',12)), sg.Input(key='FILENAME', expand_x=True)],
                  [sg.Button("Submit", font=('Calibri', 8), enable_events=True, key='SUBMIT', expand_x=False),
                   sg.Button("Exit", key='EXIT', button_color='red', expand_x=False, font=('Calibri', 8))]
              ]

    window = sg.Window('CEIC Extraction',layout,  use_default_focus=False, font=('Calibri',12))
    filename_template = 'CEIC_extraction_template'
    
    all_data = {}
    fpath = ''
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'EXIT':
            break
        if event == 'BACK_HOME':
            break
        if event == 'SUBMIT':
            if filename_template in values['Input_1']: #check if correct filename
                fpath = values['Input_1']
                df = pd.read_excel(values['Input_1'], sheet_name = 'CEIC_IDs')
                all_data = df.to_dict('list')
                event = 'SUBMIT_EXCEL'  
                break
            else:
                sg.popup('Please select the correct file template. Return to the previous screen and upload the correct file.',
                        font=('Calibri', 12), button_justification='center', title='Error Warning!')
    window.close()
    return event, all_data, values['Input_2'], values['FILENAME']


# In[11]:


def get_all_data(all_data):
    all_dfs = []
    monthly_dfs = []
    quarterly_dfs = []
    series_ids = all_data['Series ID']
    categories = all_data['Category']
    failed_ids = []
    for i in range(len(series_ids)):
        ceic_data = get_ceic_data(series_ids[i])
        if isinstance(ceic_data, dict) == False:
            failed_ids.append(series_ids[i])
        else:
            ceic_data['category'] = categories[i]
            if ceic_data['Frequency'] == 'Monthly':
                monthly_dfs.append(ceic_data)
            elif ceic_data['Frequency'] == 'Quarterly':
                quarterly_dfs.append(ceic_data)
            all_dfs.append(ceic_data)
    monthly_ls = []
    monthly_info = {'Economic Indicators':[],
                    'Freq':[],
                    'Unit':[],
                    'Source':[],
                    'Series Name':[],
                    'Compiled by':[],
                    'Category':[],
                    'include':[],
                    'diff':[],
                    'year':[],
                    'norm':[]}
    for d in monthly_dfs:
        df = d['Data']
        df.columns = ['date',d['Name']]
        monthly_ls.append(df)
        monthly_info['Economic Indicators'].append(d['Name'])
        monthly_info['Freq'].append('M')
        monthly_info['Unit'].append(d['Unit'])
        monthly_info['Source'].append(d['Source'])
        monthly_info['Series Name'].append(d['Name'])
        monthly_info['Compiled by'].append(d['Source'])
        monthly_info['Category'].append(d['category'])
        monthly_info['include'].append(1)
        monthly_info['diff'].append(0)
        monthly_info['year'].append(1)
        monthly_info['norm'].append(1)
    monthly_df = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), monthly_ls).sort_values(by=['date'], ascending=True)
    quarterly_ls = []
    quarterly_info = {'Economic Indicators':[],
                    'Freq':[],
                    'Unit':[],
                    'Source':[],
                    'Series Name':[],
                    'Compiled by':[],
                    'Category':[],
                    'include':[],
                    'diff':[],
                    'year':[],
                    'norm':[]}
    for d in quarterly_dfs:
        df = d['Data']
        df.columns = ['date',d['Name']]
        quarterly_ls.append(df)
        monthly_ls.append(df)
        quarterly_info['Economic Indicators'].append(d['Name'])
        quarterly_info['Freq'].append('Q')
        quarterly_info['Unit'].append(d['Unit'])
        quarterly_info['Source'].append(d['Source'])
        quarterly_info['Series Name'].append(d['Name'])
        quarterly_info['Compiled by'].append(d['Source'])
        quarterly_info['Category'].append(d['category'])
        quarterly_info['include'].append(1)
        quarterly_info['diff'].append(0)
        quarterly_info['year'].append(1)
        quarterly_info['norm'].append(1)
        
    try:
        quarterly_df = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), quarterly_ls).sort_values(by=['date'], ascending=True)
        quarterly_df['date'] = pd.to_datetime(quarterly_df['date'])
        quarterly_df['date'] = quarterly_df['date'].dt.strftime('%Y-%m-%d')
    except TypeError:
        quarterly_df = pd.DataFrame(columns=quarterly_info)
    
    monthly_df['date'] = pd.to_datetime(monthly_df['date'])
    monthly_df['date'] = monthly_df['date'].dt.strftime('%m/%d/%Y')
    
    return monthly_df, quarterly_df, pd.DataFrame(monthly_info), pd.DataFrame(quarterly_info), failed_ids


# In[12]:


def save_data(filepath, output_name, all_data):
    filename = filepath+'/'+'EAI_excel_template_'+output_name+'.xlsx'
    instructions_df = pd.DataFrame({'Column':['Freq',np.nan,'Include',np.nan,'Category',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,'diff',np.nan,'year',np.nan],
                                    'Accepted Values':['Q','M',0,1,'consumption','exo','financial','government','investment','trade','target_variable',0,1,0,1],
                                    'Legend':['Quarterly','Monthly','Exclude in model processing','Include in model processing',np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,np.nan,'No transformation','QoQ/MoM transformation','No transformation','YoY transformation']})
    
    monthly_df, quarterly_df, monthly_info, quarterly_info, failed_ids = get_all_data(all_data)
    
    write_excel(filename, 'Instructions', instructions_df)
    write_excel(filename, 'InfoQ', quarterly_info)
    write_excel(filename, 'QuarterlyData', quarterly_df)
    write_excel(filename, 'InfoM', monthly_info)
    write_excel(filename, 'MonthData', monthly_df)
    
    wb = load_workbook(filename)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(filename)
    return failed_ids


# In[15]:


def CEIC_extraction_process():
    event = 'START'
    while True:
        if event == 'START' or event == 'BACK_HOME':
            event = UI_home_page_design()
        if event == 'CONTINUE_HOME':
            event = UI_ceic_login()
        if event == 'OK_LOGIN':
            event, all_data, output_path, output_name = UI_add_excel_data()
        if event == 'SUBMIT_EXCEL':
            failed_ids = save_data(output_path, output_name, all_data)
            if failed_ids != []:
                failed_txt = ''
                for i in failed_ids:
                    failed_txt += str(i)+'\n'
                sg.Window('Warning', [[sg.T('The following series IDs could not be extracted from CEIC.', font=('Calibri Bold', 12))], [sg.T(failed_txt,font=('Calibri', 10))], [sg.OK(s=10)]], disable_close=True).read(close=True)
            sg.popup('Operations completed.', font=('Calibri Bold', 12))
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break


# In[16]:


CEIC_extraction_process()


# In[ ]:




