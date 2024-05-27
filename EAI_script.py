#!/usr/bin/env python
# coding: utf-8

# In[1]:


#Importing dependencies. GUI window is created to inform the user that dependencies are being imported and will close once the process is complete.

import PySimpleGUI as sg
sg.theme('DarkBlue13')

layout = [
    [sg.Text('Importing dependencies. Please wait.', font=('Calibri Bold', 12))]
]

window = sg.Window('Economic Activity Index: Import Dependencies',layout, use_default_focus=False, font=('Calibri',12))


event, values = window.read(timeout=10)
import os.path
import pandas as pd
import numpy as np
import textwrap
import time
import xlwings as xw
import pandas as pd
import datetime as dt
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import statistics
import cloudscraper
import PIL
import io
from pathlib import Path

import seaborn as sns
import plotly.express as px
import plotly
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.io as pio
import io
from PIL import Image
import os, sys
import matplotlib
from matplotlib import cm
from matplotlib import pyplot as plt
from matplotlib.patches import Circle, Wedge, Rectangle
from matplotlib.backends.backend_pdf import PdfPages
from pypdf import PdfMerger
import warnings
import calendar
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error
from sklearn.metrics import mean_absolute_percentage_error
from sklearn.metrics import r2_score

pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings("ignore", category=DeprecationWarning) 
warnings.simplefilter("ignore", UserWarning)
# !pip install openpyxl
# !pip install kaleido

import warnings
warnings.filterwarnings('ignore')
import re
import subprocess
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from openpyxl.formula.translate import Translator
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timezone
import shutil
import math
import statsmodels.api as sm
import quantecon
import statistics
from cif import cif
from functools import reduce  
import os, sys
from statsmodels.tsa.arima.model import ARIMA
from sklearn.decomposition import PCA
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import Lasso
from sklearn.ensemble import RandomForestRegressor
from sklearn.preprocessing import StandardScaler
from lightgbm import LGBMRegressor
from xgboost.sklearn import XGBRegressor
from catboost import CatBoostRegressor
from sklearn.linear_model import SGDRegressor
from sklearn.kernel_ridge import KernelRidge
from sklearn.linear_model import ElasticNet
from sklearn.linear_model import BayesianRidge
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.svm import SVR
import shap
#shap.initjs()

window.close()


# ## Home

# In[2]:


#Resize an image from the web based on the resize parameter.
def resize_img(url, resize):
    data_url = (cloudscraper.create_scraper(browser={"browser": "firefox", "platform": "windows", "mobile": False})
        .get(url)
        .content)
    img = PIL.Image.open(io.BytesIO(data_url))
    cur_width, cur_height = img.size
    new_width, new_height = resize
    scale = min(new_height / cur_height, new_width / cur_width)
    img = img.resize((int(cur_width * scale), int(cur_height * scale)), PIL.Image.ANTIALIAS)
    png_bio = io.BytesIO()
    img.save(png_bio, format="PNG")
    png_data = png_bio.getvalue()
    return png_data

#Open the home page screen of the GUI.
def UI_home_page_design():
    png_data_adb = resize_img('https://upload.wikimedia.org/wikipedia/commons/thumb/4/43/Asian_Development_Bank_logo.svg/2048px-Asian_Development_Bank_logo.svg.png', (118,118))
    png_data_banner = resize_img('https://asianbondsonline.adb.org/macroeconomictracker/images/woman-planting-with-charts-trackingasia2.jpg', (300, 300))

    intro_txt = "Developed by the Asian Development Bank for the TrackingAsia project, the Economic Activity Index enables monthly tracking of business cycles by utilizing economic series from six categories —consumption, investment, trade, government, financial, and the external sector. Further, this tool identifies the key drivers behind economic expansions and downturns over time. The goal of this tool is to provide invaluable insights into economic trends, facilitating informed decision-making on a global scale. This tool is designed to empowers users to utilize their unique datasets in constructing their own Economic Activity Index."
    welcome_txt = sg.Text('Economic Activity Index Modelling Interface', font=('Calibri Bold', 14, "bold"), expand_x=True, justification='center')
    intro_txt = sg.Text(intro_txt, size=(40, None), font=('Calibri', 12), expand_x=True, justification='center')
    continue_button = sg.Button('Continue', key='CONTINUE_HOME', font=('Calibri', 8), expand_x=True)
    exit = sg.Button('Exit', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True)
    adb_banner = sg.Image(data=png_data_adb, key='-ADB-', pad=(0,0))
    banner = sg.Image(data=png_data_banner, key='-IMAGE-', pad=(0,0))
    layout = [
        [adb_banner, banner],
        [welcome_txt],
        [intro_txt],
        [continue_button, exit]
    ]
    window = sg.Window('Economic Activity Index: Home Page', layout, grab_anywhere=True, resizable=True, font=('Calibri',12))
    while True:
        event, values = window.read()
        if event == 'CONTINUE_HOME':
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break
    window.close()
    return event


# ## Excel File Adding

# In[3]:


#Open the screen to browse the local device and select the excel file to be used in the analysis.
#Also checks if the file is formatted correctly
def UI_add_excel_data(): 
    layout = [
                  [sg.Button("Back", key='BACK_HOME', expand_x=False, font=('Calibri',8))],
                  [sg.Text(f'Select the Excel file template containing features that will be used to forecast the Economic Activity Index',
                          justification='center', expand_x=True, font=('Calibri', 12))],
                  [sg.Text('Excel File', font=('Calibri',12)), sg.Input(key='Input_1', size=(100,1)), 
                   sg.FileBrowse(font=('Calibri', 8), tooltip='Click to locate the Excel File template.')],
                  [sg.Button("Submit", font=('Calibri', 8), enable_events=True, key='SUBMIT', expand_x=False),
                   sg.Button("Exit", key='EXIT', button_color='red', expand_x=False, font=('Calibri', 8))]
              ]

    window = sg.Window('Economic Activity Index: Add Excel Data',layout,  use_default_focus=False, font=('Calibri',12))
    filename_template = 'EAI_excel_template'
    
    all_data = {}
    fpath = ''
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'EXIT':
            break
        if event == 'BACK_HOME':
            break
        if event == 'SUBMIT':
            if filename_template in values['Input_1']: #check if correct filename
                fpath = values['Input_1']
                df_qtr_info = pd.read_excel(values['Input_1'], sheet_name=1)
                df_qtr_data = pd.read_excel(values['Input_1'], sheet_name=2)
                df_month_info = pd.read_excel(values['Input_1'], sheet_name=3)
                df_month_data = pd.read_excel(values['Input_1'], sheet_name=4)
                df_info = pd.concat([df_qtr_info, df_month_info], axis=0)
                for i in df_info.values:
                    if i[1] == 'Q':
                        all_data[i[0]] = {'name': i[0],
                                         'category': i[6],
                                         'series_name': i[4],
                                         'frequency': i[1],
                                         'include': i[7],
                                         'diff': i[8],
                                         'year': i[9],
                                         'norm': i[10],
                                         'edited': 0, 
                                         'Data': df_qtr_data[['date', i[0]]]}
                    elif i[1] == 'M':
                        all_data[i[0]] = {'name': i[0],
                                         'category': i[6],
                                         'series_name': i[4],
                                         'frequency': i[1],
                                         'include': i[7],
                                         'diff': i[8],
                                         'year': i[9],
                                         'norm': i[10],
                                         'edited': 0, 
                                         'Data': df_month_data[['date', i[0]]]}
                event = 'SUBMIT_EXCEL'
                
                if 'target_variable' not in df_info.Category.values:
                    event = 'BACK_CHECK_EXCEL'
                    sg.popup("No target variable found in the Excel File, please make sure to include one target variable.",
                             font=('Calibri', 12), button_justification='center', title='Error Warning!')
                    
                if (df_info.Category == 'target_variable').sum() > 1:
                    event = 'BACK_CHECK_EXCEL'
                    sg.popup("More than 1 target variable found in the Excel File, please make sure to only include one target variable.",
                             font=('Calibri', 12), button_justification='center', title= 'Error Warning!')
                
                break
            else:
                sg.popup('Please select the correct file template. Return to the previous screen and upload the correct file.',
                        font=('Calibri', 12), button_justification='center', title='Error Warning!')
    window.close()
    return event, all_data, fpath


# ## Indicator Selection

# In[4]:


#Open UI screen that shows all indicators in the chosen file within their respective economic buckets
def UI_choose_indicators(all_data={}): 

    back_button = sg.Button('Back', key='BACK_CHECK_EXCEL', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Select indicators to use for the model(s)', font=('Calibri Bold', 14), expand_x=True, justification='center')
    instructions_txt = sg.Text("Select the indicators that you'd like to be incorporated into the model(s).", font=('Calibri', 12), expand_x=True, justification='center')
    
    country_contributions = {'country': [],
                                 'target_variable': [],
                                 'consumption': [],
                                 'exo': [],
                                 'financial': [],
                                 'government': [],
                                 'investment': [],
                                 'trade': []}

    inspect_data_button = sg.Button('   Inspect Data    ', key='INSPECT_INDICATORS', font=('Calibri', 8), expand_x=True, tooltip='Click to view visualizations and summary statistics on the indicator(s).')
    reconfigure_data_button = sg.Button(' Edit Processing ', key='RECONFIGURE_INDICATORS', font=('Calibri', 8), expand_x=True, tooltip='Click to view / edit / overwrite the inputs in the Excel File.')
    
    ok_button = sg.Button('Ok', key='OK_CHOOSE_INDICATORS', font=('Calibri', 8), expand_x=True, auto_size_button=False)
    exit = sg.Button('  Exit  ', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True, auto_size_button=False)

    layout = [
        [back_button],
        [title_txt],
        [instructions_txt],
    ]

    column_layout = []
    if all_data != {}:
        for key in all_data.keys():
            country_contributions[all_data[key]['category']].append(all_data[key]['name'])


    for key in country_contributions.keys():
        new_ls = country_contributions[key]
        new_ls_2 = []
        for l in new_ls:
            if all_data[l]['include'] == 1:
                new_ls_2.append(l)
        #for i in range(len(new_ls)):
         #   if all_data[new_ls[i]]['include'] == 0:
          #      new_ls[i] = sg.Checkbox(textwrap.fill(new_ls[i],25), default=False, key = 'contributions-->'+key+'-->'+new_ls[i], font=('Calibri', 6))
           # else:
            #    new_ls[i] = sg.Checkbox(textwrap.fill(new_ls[i],25), default=True, key = 'contributions-->'+key+'-->'+new_ls[i], font=('Calibri', 6))
        if key != 'country':
            column_layout.append([sg.Text(key.capitalize()+' Indicators:', key='-OUT-', font=('Calibri Bold', 8), expand_x=True, justification='left')])
            #column_layout.append([sg.Column([new_ls])])
            column_layout.append([sg.Listbox(new_ls_2, key='LISTBOX', font=('Calibri', 8), expand_x=True,pad=(0,0),expand_y=True,background_color=sg.theme_background_color(),text_color='white')])

    layout.append([[sg.Column(column_layout, scrollable=True, expand_x=True)],[sg.Push(),inspect_data_button, reconfigure_data_button, sg.Push()],[ok_button,exit]])

    window = sg.Window('Economic Activity Index: Indicator Selection', layout, grab_anywhere=True,resizable=True, font=('Calibri',12))
    while True:
        event, values = window.read()
        if event == 'OK_CHOOSE_INDICATORS':
            break
        if event == 'INSPECT_INDICATORS':
            break
        if event == 'RECONFIGURE_INDICATORS':
            break
        elif event == 'BACK_CHECK_EXCEL':
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break
    
    window.close()
    return event, values, all_data


# ## Model Selection

# In[5]:


SYMBOL_UP =    '▲'
SYMBOL_DOWN =  '▼'

model_list = ['Linear Regression','Random Forest','Light Gradient Boosting Method (LGBM)','Extreme Boosting (XGBoost)','CatBoost','Stochastic Gradient Descent (SGDR)','Kernel Ridge','Elastic Net','Bayesian Ridge','Gradient Boosting','Support Vector Machine (SVM)']

models_explained_text = """Linear regression: 
Plots a line of best fit for each type of linear regression using the least squares method.

Random Forest: 
Uses ensemble learning with multiple decision trees to find more accurate predictions.

Light Gradient Boosting Method (LGBM): 
More efficient and lightweight gradient boosting method.

Extreme Boosting (XGBoost): 
Uses advanced regularization to reduce overfitting. Usually performs better with larger datasets.

CatBoost: 
Builds upon the theory of decision trees and gradient boosting by growing oblivious trees to simplify the model and improve efficiency. 

Stochastic Gradient Descent (SGDR): 
Iterative method that optimizes an objective function and reduces the computational burden of bigger datasets.

Kernel Ridge: 
Uses a kernel function to calculate weights. This model is used when there is too much data for a traditional linear model.

Elastic Net: 
Uses lasso and ridge techniques to improve the regularization of models.

Bayesian Ridge: 
Uses linear regression with probability distributors instead of point estimates.

Gradient Boosting: 
An ensemble of weak learners/models are used to create a single, more accurate model.

Support Vector Machine (SVM): 
Finds the hyperplane that passes through as many data points as possible within a certain distance to minimize prediction error.
 """

#Creates a collapsible column element within a UI screen
def Collapsible(layout, key, title='', arrows=(sg.SYMBOL_DOWN, sg.SYMBOL_UP), collapsed=False):
    """
    User Defined Element
    A "collapsable section" element. Like a container element that can be collapsed and brought back
    :param layout:Tuple[List[sg.Element]]: The layout for the section
    :param key:Any: Key used to make this section visible / invisible
    :param title:str: Title to show next to arrow
    :param arrows:Tuple[str, str]: The strings to use to show the section is (Open, Closed).
    :param collapsed:bool: If True, then the section begins in a collapsed state
    :return:sg.Column: Column including the arrows, title and the layout that is pinned
    """
    return sg.Column([[sg.pin(sg.Column(layout, key=key, visible=not collapsed, metadata=arrows,scrollable=True))]], pad=(0,0))

#Opens the model selection screen in the GUI
def UI_select_model():
    
    back_button = sg.Button('Back', key='BACK_SELECT_INDICATORS', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Select the model(s) to be used:', font=('Calibri Bold', 14), expand_x=True, justification='center')
    instructions_txt = sg.Text('Select the machine learning model(s) to run and the corresponding start date for training the model(s).\nBy default, 50% of the dataset will be initialized as the training dataset unless otherwise stated below.', 
                               font=('Calibri', 12), expand_x=True, justification='center')
    model_explanations = sg.Text(models_explained_text, key='-OUT-', font=('Calibri', 8), expand_x=True, justification='left')

    checkbox_ls = []
    for m in model_list:
        checkbox_ls.append([sg.Checkbox(m, key = m, font=('Calibri', 8), expand_x = True,default=False)])

    model_select = sg.Column(checkbox_ls, scrollable=True, key='MODEL_SELECT', expand_x=True)
    ok_button = sg.Button('Ok', key='OK_SELECT_MODELS', font=('Calibri', 8), expand_x=True)
    exit = sg.Button('Exit', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True)
    month_selector_start = sg.Combo(['Jan','Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep',
                              'Oct', 'Nov', 'Dec'], enable_events = True, default_value=None, key='MONTH_SELECT_START', font=('Calibri', 8), expand_x=False)
    month_selector_end = sg.Combo(['Jan','Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep',
                               'Oct', 'Nov', 'Dec'], enable_events = True, default_value=None, key='MONTH_SELECT_END', font=('Calibri', 8), expand_x=False)
    year_selector_start = sg.Combo(list(range(1990, dt.datetime.today().year+1, 1)), enable_events = True, default_value=None, key='YEAR_SELECT_START', font=('Calibri', 8), expand_x=False)
    year_selector_end = sg.Combo(list(range(1990, dt.datetime.today().year+1, 1)), enable_events = True, default_value=None, key='YEAR_SELECT_END', font=('Calibri', 8), expand_x=False)
    
    output_path = [sg.Text("Output Folder:", font=('Calibri Bold', 8)), sg.Input(key='-INPUT-'), sg.FolderBrowse(target='-INPUT-', font=('Calibri', 8), key='OUTPUT_PATH', tooltip='Click to select the destination folder for all data outputs.')] # Target with key '-INPUT-'

    section1 = [[model_explanations]]
    
    layout = [
        [back_button],
        [title_txt],
        [instructions_txt],
        [sg.Text("Start date of training:", font=('Calibri Bold', 8)), month_selector_start, year_selector_start, sg.Text("End date of training:", font=('Calibri Bold', 8)), month_selector_end, year_selector_end],
        [output_path],
        [sg.T(SYMBOL_UP, enable_events=True, k='-OPEN SEC1-', text_color='white'), sg.T('Open/Close Model Guide', font = ('Calibri', 8), enable_events=True, text_color='white', k='-OPEN SEC1-TEXT-')],
        [Collapsible(section1, '-SEC1-', collapsed=True)],
        [model_select],
        [ok_button, exit]
    ]
    window = sg.Window('Economic Activity Index: Select Model(s)', layout, grab_anywhere=True,resizable=True, font=('Calibri',12))
    opened1 = False
    date_dict = {}
    output_path = ''
    while True:
        event, values = window.read()
        if event == '-OPEN SEC1-':
            opened1 = not opened1
            window['-OPEN SEC1-'].update(SYMBOL_DOWN if opened1 else SYMBOL_UP)
            window['-SEC1-'].update(visible=opened1)
        if event == 'OK_SELECT_MODELS':
            if all(v == False for v in values.values()):
                sg.Popup('Please select at least one model to use for this analysis.', font=('Calibri', 12), grab_anywhere=True, button_justification='center', title='Error Warning!')
            if values['OUTPUT_PATH'] == '':
                sg.Popup('Please indicate an output filepath for the generated files.', font=('Calibri', 12), grab_anywhere=True, button_justification='center', title='Error Warning!')
            else:
                try:
                    date_dict['start_date'] = dt.datetime.strptime(values['MONTH_SELECT_START'] + str(values['YEAR_SELECT_START']), '%b%Y').strftime('%Y-%m-%d')
                except:
                    date_dict['start_date'] = ''
                try:
                    date_dict['end_date'] = dt.datetime.strptime(values['MONTH_SELECT_END'] + str(values['YEAR_SELECT_END']), '%b%Y').strftime('%Y-%m-%d')
                except:
                    date_dict['end_date'] = ''
                output_path = values['OUTPUT_PATH']
                break
        elif event == 'BACK_SELECT_INDICATORS':
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break
    window.close()
    return event, values, date_dict, output_path 


# ## Data Inspection

# In[6]:


#Helper function used to draw figures for plotting charts
def draw_figure(canvas, figure, loc=(0, 0)):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='top', fill='both', expand=1)
    return figure_canvas_agg

#Helper function to delete drawn figures
def delete_figure_agg(figure_agg):
    figure_agg.get_tk_widget().forget()
    plt.close('all')


# In[7]:


#Get the names of all indicator series in the excel file
def get_all_series_names(fpath, sheet):
    
    #Monthly sheet: MonthData
    #Quarterly sheet: QuarterlyData
    
    df = pd.read_excel(fpath, sheet_name = sheet)
    return list(df.columns)[1:]

#Retrieve a particular series as a pandas dataframe
def get_series(fpath, series_name, sheet):
    
    #Monthly sheet: MonthData
    #Quarterly sheet: QuarterlyData
    
    df = pd.read_excel(fpath, sheet_name = sheet)
    df = df[['date', series_name]].dropna()
    try:
        df['date'] = pd.to_datetime(df['date'], unit='D', origin='1899-12-30')
        df['date'] = df['date'].apply(lambda dt: dt.replace(day=1))
    except:
        pass
    
    df['date'] = pd.to_datetime(df['date'])
    
    return df


# In[8]:


#Get information on each indicator series with regards to their particular processing requirements
def get_series_information(fpath, series_name, sheet):
    if sheet == 'MonthData':
        info_sheet = 'InfoM'
    else:
        info_sheet = 'InfoQ'
    df = pd.read_excel(fpath, sheet_name = info_sheet)
    df = df.T
    new_header = df.iloc[0]
    df = df[1:]
    df.columns = new_header
    new_dict = df.to_dict()
    return new_dict[series_name]


# In[9]:


#Popup screen when saving a chart
def save_chart_popup(series_name):
    title_text = sg.Text('Save Series Chart: '+ series_name, font=('Calibri Bold', 14), expand_x=True, justification='center')
    browse_input = sg.Input('Filepath...', key = 'FILEPATH', disabled = True, font=('Calibri',12))
    browse_button = sg.FileSaveAs('Browse', enable_events=True, key='BROWSE', font=('Calibri', 8), expand_x=True, file_types = (('PNG','.png'),('JPEG','.jpg'),('PDF','.pdf'),), tooltip='Input the filename and select the destination folder for the chart.')
    save_button = sg.Button('Save', key='SAVE_AS', font=('Calibri', 8), expand_x=True)
    cancel_button = sg.Button('Cancel', key='CANCEL', font=('Calibri', 8), button_color='red', expand_x=True)
    layout = [
        [title_text],
        [browse_input, browse_button],
        [save_button, cancel_button]
    ]
    window = sg.Window('Economic Activity Index: Save Series Chart', layout, finalize = True, grab_anywhere=True,resizable=True, font=('Calibri',12))
    while True:
        event, values = window.read()
        if event == 'SAVE_AS':
            if values['BROWSE'] == '':
                sg.Popup('Please indicate a filepath before saving.')
            else:
                break
        elif event == 'CANCEL' or event == WIN.CLOSED:
            break
    window.close()
    return event, values

#Popup screen when saving data
def save_data_popup(series_name):
    title_text = sg.Text('Save Series Dataset: '+ series_name, font=('Calibri Bold', 14), expand_x=True, justification='center')
    browse_input = sg.Input('Filepath...', key = 'FILEPATH', disabled = True, font=('Calibri',12))
    browse_button = sg.FileSaveAs('Browse', enable_events=True, key='BROWSE', font=('Calibri', 8), expand_x=True, file_types = (('CSV','.csv'),), tooltip='Input the filename and select the destination folder for the data series.')
    save_button = sg.Button('Save', key='SAVE_AS', font=('Calibri', 8), expand_x=True)
    cancel_button = sg.Button('Cancel', key='CANCEL', font=('Calibri', 8), button_color='red', expand_x=True)
    layout = [
        [title_text],
        [browse_input, browse_button],
        [save_button, cancel_button]
    ]
    window = sg.Window('Economic Activity Index: Save Series Data', layout, finalize = True, grab_anywhere=True,resizable=True, font=('Calibri',12))
    while True:
        event, values = window.read()
        if event == 'SAVE_AS':
            if values['BROWSE'] == '':
                sg.Popup('Please indicate a filepath before saving.')
            else:
                break
        elif event == 'CANCEL' or event == WIN.CLOSED:
            break
    window.close()
    return event, values


# In[10]:


#Open inspect processed data GUI screen. Allows the user to inspect a data series after processing the data based on their selections.
def UI_inspect_processed_data(df, info_dict):
    
    ### Process dataframe
    
    series_name = info_dict['SERIES_SELECT']
    processing_txt_ls = []
    
    if info_dict['DIFF'] == True:
        processing_txt_ls.append(sg.Text('- Month on Month Differential', justification = 'left', font=('Calibri',8)))
        if info_dict['FREQUENCY_SELECT'] == 'Monthly':
            df[series_name] = df[series_name].pct_change(1)
        else:
            pass
    if info_dict['YEAR'] == True:
        processing_txt_ls.append(sg.Text('- Year on Year Growth', justification = 'left', font=('Calibri',8)))
        if info_dict['FREQUENCY_SELECT'] == 'Monthly':
            df[series_name] = df[series_name].pct_change(12)
        else:
            df[series_name] = df[series_name].pct_change(4)
        df = df.dropna()
    if info_dict['NORM'] == True:
        processing_txt_ls.append(sg.Text('- Normalize', justification = 'left', font=('Calibri',8)))
        new_df = df[df.columns[1:]]
        normalized_df=(new_df-new_df.mean())/new_df.std()
        normalized_df.insert(0, 'date', df['date'])
        df = normalized_df.copy()
    
    if processing_txt_ls == []:
        processing_txt_ls.append(sg.Text('No Change in Data', justification = 'left', font = ('Calibri',8)))
    ###
    
    back_button = sg.Button('Back', key='BACK_INSPECT_DATA', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Post Processed Data Inspection: '+series_name, font=('Calibri Bold', 14), expand_x=True, justification='center')
    intro_txt = sg.Text('The following chart was derived after processing the original dataset based on your chosen processing steps.', font=('Calibri', 12), expand_x=True, justification='center')
    download_chart_button = sg.Button('Save Chart', enable_events=True, key='SAVE_CHART', font=('Calibri', 8), expand_x=True)
    save_data_button = sg.Button('Save Data', enable_events=True, key='SAVE_DATA', font=('Calibri', 8), expand_x=True)
    
    plt.plot(list(df['date']), list(df[series_name]), color='blue', linewidth=0.5)
    plt.title(series_name+ ' Processed', fontsize = 8)
    plt.xlabel('Date', fontsize=6)
    plt.ylabel(series_name, fontsize=6)
    plt.yticks(fontsize=6)
    plt.xticks(fontsize=6)
    plt.grid(True)
    fig = plt.gcf()      # if using Pyplot then get the figure from the plot
    figure_x, figure_y, figure_w, figure_h = fig.bbox.bounds
    
    column_layout2 = [[sg.Canvas(size=(figure_w, figure_h), key='canvas')],
                      [save_data_button, download_chart_button]]
    
    describe_dict = df.describe().to_dict()
    table_elements = [
        ['First Available Date', str(describe_dict['date']['min'].date())],
        ['Latest Available Date', str(describe_dict['date']['max'].date())],
        ['Value Count', str(round(describe_dict[series_name]['count'],2))],
        ['Mean', str(round(describe_dict[series_name]['mean'],2))],
        ['Minimum', str(round(describe_dict[series_name]['min'],2))],
        ['25th Percentile', str(round(describe_dict[series_name]['25%'],2))],
        ['50th Percentile', str(round(describe_dict[series_name]['50%'],2))],
        ['75th Percentile', str(round(describe_dict[series_name]['75%'],2))],
        ['Maximum', str(round(describe_dict[series_name]['max'],2))],
        ['Standard Deviation', str(round(describe_dict[series_name]['std'],2))],
    ]
    table_headers = ['Statistic', series_name]
    
    
    column_layout3 = [
        [sg.Text('Summary Statistics', font=('Calibri Bold',10))],
        [sg.Table(table_elements, headings=table_headers, cols_justification=['c','c'], auto_size_columns=True, expand_y=True, font=('Calibri',8), header_font=('Calibri Bold',10))],
        [sg.Text('Selected Processing Steps', font=('Calibri Bold',10))]
    ]
    for t in processing_txt_ls:
        column_layout3.append([t])
    
    layout = [
        [back_button],
        [title_txt],
        [intro_txt],
        [sg.Column(column_layout2, element_justification='center'), sg.Column(column_layout3, element_justification='center')],
    ]
    
    window = sg.Window('Economic Activity Index: Inspect Data', layout, location = (0,0), finalize = True, grab_anywhere=True,resizable=True)
    fig_canvas_agg = draw_figure(window['canvas'].TKCanvas, fig)
    
    while True:
        event, values = window.read()
        if event == 'BACK_INDICATOR_SELECT':
            break
        if event == 'SAVE_DATA':
            save_event, save_values = save_data_popup(series_name)
            if save_event == 'SAVE_AS':
                fpath = save_values['BROWSE']
                try:
                    df.to_csv(fpath)
                except:
                    df.to_excel(fpath)
        if event == 'SAVE_CHART':
            save_event, save_values = save_chart_popup(series_name)
            if save_event == 'SAVE_AS':
                plt.savefig(save_values['BROWSE'])
        elif event == 'FREQUENCY_SELECT' or event == 'SERIES_SELECT':
            break
        elif event == sg.WIN_CLOSED or event == 'BACK_INSPECT_DATA':
            event = 'BACK_INSPECT_DATA'
            break
    plt.close()
    window.close()
    return event, values


# In[11]:


#Opens GUI screen to inspect a particular data series
def UI_inspect_data(fpath, series_name, current_frequency):

    if current_frequency == 'Monthly':
        sheet = 'MonthData'
    else:
        sheet = 'QuarterlyData'
    
    if series_name == '':
        series_name = get_all_series_names(fpath, sheet)[0]
        df = get_series(fpath, series_name, sheet)
        default_info = get_series_information(fpath, series_name, sheet)
    elif series_name not in get_all_series_names(fpath, sheet):
        series_name = get_all_series_names(fpath, sheet)[0]
        df = get_series(fpath, series_name, sheet)
        default_info = get_series_information(fpath, series_name, sheet)
    else:
        df = get_series(fpath, series_name, sheet)
        default_info = get_series_information(fpath, series_name, sheet)
    new_dict = default_info.copy()
    
    back_button = sg.Button('Back', key='BACK_INDICATOR_SELECT', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Data Inspection: '+series_name, font=('Calibri Bold', 14), expand_x=True, justification='center')
    intro_txt = sg.Text('Use this screen to inspect a selected indicator from the provided data file. \nYou may also alter the pre-processing procedures for this indicator prior to running the machine learning framework.', font=('Calibri', 12), expand_x=True, justification='center')
    download_chart_button = sg.Button('Save Chart', enable_events=True, key='SAVE_CHART', font=('Calibri', 8), expand_x=True)
    save_data_button = sg.Button('Save Data', enable_events=True, key='SAVE_DATA', font=('Calibri', 8), expand_x=True)
    #download_chart_input = sg.Input(key='SAVE_CHART_INPUT')
    exit = sg.Button('Exit', key='EXIT', font=('Calibri', 8), button_color='red', expand_x=True)
    
    plt.plot(list(df['date']), list(df[series_name]), color='blue', linewidth=0.5)
    plt.title(series_name+ ' ('+sheet+')', fontsize = 8)
    plt.xlabel('Date', fontsize=6)
    plt.ylabel(series_name, fontsize=6)
    plt.yticks(fontsize=6)
    #f = pd.date_range(list(df['date'])[0], list(df['date'])[-1],20) #these will be x-axis labels
    #plt.xticks(f, fontsize=6, rotation=45)
    plt.xticks(fontsize=6)
    plt.grid(True)
    #plt.figtext(0.1,0.5, df.describe().to_string())
    fig = plt.gcf()      # if using Pyplot then get the figure from the plot
    figure_x, figure_y, figure_w, figure_h = fig.bbox.bounds
    
    frequency_selector = sg.Combo(['Monthly','Quarterly'], enable_events = True, default_value=current_frequency, key='FREQUENCY_SELECT', font=('Calibri', 8), expand_x=True, size=(15,1))
    series_selector = sg.Combo(get_all_series_names(fpath, sheet), enable_events = True, default_value = series_name, key = 'SERIES_SELECT', font=('Calibri', 8), expand_x=True, size=(15,1))
    category_selector = sg.Combo(['target_variable','consumption','exo','financial','government','investment','trade'], default_value = default_info['Category'], key='CATEGORY', font=('Calibri', 8), expand_x=True, size=(15,1))
    include_bool = True if default_info['include'] == 1 else False
    include_check = sg.Checkbox('Include in Model(s)', default=include_bool, key='INCLUDE', font=('Calibri',8))
    diff_bool = True if default_info['diff'] == 1 else False
    diff_check = sg.Checkbox('Month on Month Differential', default=diff_bool, key='DIFF', font=('Calibri',8))
    year_bool = True if default_info['year'] == 1 else False
    year_check = sg.Checkbox('Year on Year Change', default=year_bool, key='YEAR', font=('Calibri',8))
    norm_bool = True if default_info['norm'] == 1 else False
    norm_check = sg.Checkbox('Normalize', default=norm_bool, key='NORM', font=('Calibri',8))
    inspect_processed_button = sg.Button('Inspect Processed Data', enable_events=True, key='INSPECT_PROCESSED_DATA', font=('Calibri', 8), expand_x=True, tooltip="Select to view the visualization and summary statistics of the processed data based on the pre-processing steps selected.")
    
    column_layout1 = [
                      [sg.Text('Data Options', font=('Calibri Bold',10), justification='center')],
                      [sg.Text('Select Frequency:', font = ('Calibri',8), justification='left'), frequency_selector],
                      [sg.Text('Select Series:', font = ('Calibri',8), justification='left'), series_selector],
                      [sg.Text('Data Category:', font = ('Calibri',8), justification='left'), category_selector],
                      [include_check],
                      [diff_check],
                      [year_check],
                      [norm_check],
                      [inspect_processed_button]
                ]
    
    
    column_layout2 = [[sg.Canvas(size=(figure_w, figure_h), key='canvas')],
                      [save_data_button, download_chart_button]]
    
    describe_dict = df.describe().to_dict()
    table_elements = [
        ['First Available Date', str(describe_dict['date']['min'].date())],
        ['Latest Available Date', str(describe_dict['date']['max'].date())],
        ['Value Count', str(round(describe_dict[series_name]['count'],2))],
        ['Mean', str(round(describe_dict[series_name]['mean'],2))],
        ['Minimum', str(round(describe_dict[series_name]['min'],2))],
        ['25th Percentile', str(round(describe_dict[series_name]['25%'],2))],
        ['50th Percentile', str(round(describe_dict[series_name]['50%'],2))],
        ['75th Percentile', str(round(describe_dict[series_name]['75%'],2))],
        ['Maximum', str(round(describe_dict[series_name]['max'],2))],
        ['Standard Deviation', str(round(describe_dict[series_name]['std'],2))],
    ]
    table_headers = ['Statistic', series_name]
    column_layout3 = [
        [sg.Text('Summary Statistics', font=('Calibri Bold',10))],
        [sg.Table(table_elements, headings=table_headers, cols_justification=['c','c'], auto_size_columns=True, expand_y=True, font=('Calibri',8), header_font=('Calibri Bold',10))]
    ]
    
    
    layout = [
        [back_button],
        [title_txt],
        [intro_txt],
        [sg.Column(column_layout1, element_justification='left'), sg.Column(column_layout2, element_justification='center'), sg.Column(column_layout3, element_justification='center')],
        [exit]
    ]
    
    window = sg.Window('Economic Activity Index: Inspect Data', layout, location = (0,0), finalize = True, grab_anywhere=True,resizable=True)
    fig_canvas_agg = draw_figure(window['canvas'].TKCanvas, fig)
    
    while True:
        event, values = window.read()
        if event == 'BACK_INDICATOR_SELECT':
            break
        if event == 'INSPECT_PROCESSED_DATA':
            break
            new_event, new_values = UI_inspect_processed_data(df, values)
        if event == 'SAVE_DATA':
            save_event, save_values = save_data_popup(series_name)
            if save_event == 'SAVE_AS':
                fpath = save_values['BROWSE']
                try:
                    df.to_csv(fpath)
                except:
                    df.to_excel(fpath)
        if event == 'SAVE_CHART':
            save_event, save_values = save_chart_popup(series_name)
            if save_event == 'SAVE_AS':
                plt.savefig(save_values['BROWSE'])
        elif event == 'FREQUENCY_SELECT' or event == 'SERIES_SELECT':
            break
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            break
    plt.close()
    window.close()
    try:
        new_dict['Category'] = values['CATEGORY']
        new_dict['include'] = 1 if values['INCLUDE'] == True else 0
        new_dict['diff'] = 1 if values['DIFF'] == True else 0
        new_dict['year'] = 1 if values['YEAR'] == True else 0
        new_dict['norm'] = 1 if values['NORM'] == True else 0

        if default_info == new_dict:
            default_info['edited'] = 0
        else:
            default_info['edited'] = 1
    except:
        default_info['edited'] = 0
    
    return event, values, default_info, df


# In[12]:


#Opens the GUI screen for the entire inspect data flow. Utilizes all inspect data functions.
def inspect_data_flow(fpath):
    event, values, info_dict, df = UI_inspect_data(fpath, '', 'Monthly')
    while True:
        if event == 'SERIES_SELECT' or event == 'FREQUENCY_SELECT' or event == 'BACK_INSPECT_DATA':
            event, values, info_dict, df = UI_inspect_data(fpath, values['SERIES_SELECT'], values['FREQUENCY_SELECT'])
        if event == 'INSPECT_PROCESSED_DATA':
            event, _ = UI_inspect_processed_data(df, values)
        elif event == 'BACK_INDICATOR_SELECT':
            event = 'BACK_SELECT_INDICATORS'
            break
        elif event == 'EXIT' or event == None:
            event = 'BACK_SELECT_INDICATORS'
            break
    return event, values, info_dict


# ## Reconfigure Data Processing

# In[13]:


#Get data from a dictionary of values from the monthly and quarterly information sheets of the excel file.
def get_data_from_table(table_values):
    key_ls = list(table_values.keys())
    new_dict = {'Economic Indicators':[], 
                'Freq':[], 
                'Unit':[], 
                'Source':[], 
                'Series Name':[], 
                'Compiled by':[], 
                'Category':[], 
                'include':[], 
                'diff':[], 
                'year':[],
                 'norm': []} #added
    
    for key in list(table_values.keys())[:-1]:
        new_dict[key[:-2]].append(table_values[key])
        
    return pd.DataFrame(new_dict)


# In[14]:


#Creates an editable version of the information sheets of the excel on the GUI. Allows users to manipulate the values in that table.
def create_editable_table(fpath, frequency):
    if frequency == 'Monthly':
        sheet = 'InfoM'
    else:
        sheet = 'InfoQ'
    df = pd.read_excel(fpath, sheet_name = sheet)
    
    cols = list(df.columns)
    headers = []
    for c in cols:
        headers.append(sg.Text(c, pad=(0,0), size=(17,1), font=('Calibri Bold', 8), justification='c'))
    
    back_button = sg.Button('Back', key='BACK_SELECT_INDICATORS', font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Edit ' + frequency +' Indicator Information', font=('Calibri Bold', 14), expand_x=True, justification='center')
    
    layout = [[back_button],[title_txt],[headers]]
    combo_keys = []
    
    for row in range(0, len(df)):
        layout.append([
            sg.Input(df.iloc[row,0], size=(17,1),pad=(0, 0), key='Economic Indicators '+str(row), font=('Calibri',8), disabled=True, justification='center'), 
            sg.Input(df.iloc[row,1], size=(17,1),pad=(0, 0), key='Freq '+str(row), font=('Calibri',8), disabled=True, justification='center'), 
            sg.Input(df.iloc[row,2], size=(17,1),pad=(0, 0), key='Unit '+str(row), font=('Calibri',8), disabled=True, justification='center'),
            sg.Input(df.iloc[row,3], size=(17,1),pad=(0, 0), key='Source '+str(row), font=('Calibri',8), disabled=True, justification='center'), 
            sg.Input(df.iloc[row,4], size=(17,1),pad=(0, 0), key='Series Name '+str(row), font=('Calibri',8), disabled=True, justification='center'),
            sg.Input(df.iloc[row,5], size=(17,1),pad=(0, 0), key='Compiled by '+str(row), font=('Calibri',8), disabled=True, justification='center'), 
            sg.Combo(['target_variable','consumption','exo','financial','government','investment','trade'], default_value= df.iloc[row,6], size=(16,1),pad=(0, 0), key='Category '+str(row), font=('Calibri',8)),
            sg.Combo([1,0], default_value=int(df.iloc[row,7]), size=(15,1),pad=(0, 0), key='include '+str(row), font=('Calibri',8)),
            sg.Combo([1,0], default_value=int(df.iloc[row,8]), size=(15,1),pad=(0, 0), key='diff '+str(row), font=('Calibri',8)),
            sg.Combo([1,0], default_value=int(df.iloc[row,9]), size=(15,1),pad=(0, 0), key='year '+str(row), font=('Calibri',8)),
            sg.Combo([1,0], default_value=int(df.iloc[row,10]), size=(15,1),pad=(0, 0), key='norm '+str(row), font=('Calibri',8))
        ])
        combo_keys = combo_keys + ['Category '+str(row),'include '+str(row),'diff '+str(row),'year '+str(row), 'norm '+str(row)]
        
    layout.append([sg.Button("Submit", key = 'SUBMIT_SPREADSHEET', font=('Calibri',8)),
                   sg.Button("Revert to Original", key = 'REVERT_ORIGINAL', font=('Calibri',8), tooltip="Select to revert back to the original file's inputs"),
                   sg.Checkbox('Save Changes to Working File', key = 'SAVE_CHANGES', font=('Calibri', 8), expand_x = False,default=False, tooltip = 'Select to overwrite the modifications made to the Excel Template.')])
    window = sg.Window('Spreadsheet', layout, font=('Calibri',12), resizable=True, grab_anywhere=True,finalize=True)
    for combo in combo_keys:
        window[combo].Widget.configure(justify='center')


    while True:
        event, values = window.read()
        if event in (sg.WIN_CLOSED, 'EXIT'):
            event = 'BACK_SELECT_INDICATORS'
            break
        if event == 'BACK_SELECT_INDICATORS':
            break
        elif event == 'REVERT_ORIGINAL':
            break
        elif event == 'SUBMIT_SPREADSHEET':
            new_table = get_data_from_table(values)
            window.close()
            event = 'BACK_SELECT_INDICATORS'
            if new_table.equals(df):
                return event, values, df
            else:
                if values['SAVE_CHANGES'] == True:
                    with pd.ExcelWriter(fpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer: #added
                        new_table.to_excel(writer, sheet_name=sheet, index=False)
                return event, values, new_table
        
    window.close()
    return event, values, df


# In[15]:


#Popup to allow users to choose which frequency information they would like to inspect.
def frequency_popup():
    choice, _ = sg.Window('Economic Activity Index: Choose Data Frequency', [[sg.T("Choose a Data Frequency to Edit", font=('Calibri Bold',12), justification='center')], [sg.Button('Monthly', font=('Calibri',8), expand_x=True), sg.Button('Quarterly', font=('Calibri',8), expand_x=True)]], disable_close=False).read(close=True)
    return choice


# ## Final Flow

# In[16]:


#Final process flow utilizing all previous functions. This function is ran before any back end functions are used.
def UI_process_flow():
    event = 'START'
    model_selection = {}
    added_data_dict = {}
    added_data = {}
    fpath = ''
    monthly_info = ''
    quarterly_info = ''
    frequency = 'Monthly'
    sg.theme('DarkBlue13')
    
    while True:
        if event == 'START' or event == 'BACK_HOME':
            event = UI_home_page_design()
        if event == 'CONTINUE_HOME' or event == 'BACK_CHECK_EXCEL':
            event, all_data, fpath = UI_add_excel_data()
            if fpath != '':
                monthly_info = pd.read_excel(fpath, sheet_name = 'InfoM')
                quarterly_info = pd.read_excel(fpath, sheet_name = 'InfoQ')
        if event == 'SUBMIT_EXCEL' or event == 'BACK_SELECT_INDICATORS':
            event, values, added_data = UI_choose_indicators(all_data)
            added_data_dict = added_data 
        if event == 'INSPECT_INDICATORS':
            event, values, info_dict = inspect_data_flow(fpath)
        if event == 'RECONFIGURE_INDICATORS':
            frequency = frequency_popup()
            if frequency == None:
                frequency = 'Monthly'
            event, values, df = create_editable_table(fpath, frequency)
            if frequency == 'Monthly':
                monthly_info = df.copy()
            else:
                quarterly_info = df.copy()
        if event == 'REVERT_ORIGINAL':
            event, values, df = create_editable_table(fpath, frequency)
        if event == 'OK_CHOOSE_INDICATORS':
            event, model_selection, date_dict, output_path = UI_select_model()
        elif event == 'OK_SELECT_MODELS':
            return event, model_selection, added_data_dict, date_dict, monthly_info, quarterly_info, output_path
        elif event == sg.WIN_CLOSED or event == 'EXIT':
            return {}, {}, {}, {}, {}, {}, {}


# # Back end Integration

# ### Helper functions

# In[17]:


#helper classes and functions
#The hidden prints class hides any string prints from a function

class HiddenPrints:
    def __enter__(self):
        self._original_stdout = sys.stdout
        sys.stdout = open(os.devnull, 'w')

    def __exit__(self, exc_type, exc_val, exc_tb):
        sys.stdout.close()
        sys.stdout = self._original_stdout


# In[18]:


#Use data outputs from UI to organize the information for each series

def separate_indicator_target_data(added_data_dict, monthly_info, quarterly_info):
    var_dict = {
                 'target_variable': {},
                 'consumption': {},
                 'exo': {},
                 'financial': {},
                 'government': {},
                 'investment': {},
                 'trade': {}}
    q_d = quarterly_info.T.to_dict()
    m_d = monthly_info.T.to_dict()
    for key in q_d.keys():
        try:
            new_dict = {}
            new_dict['frequency'] = q_d[key]['Freq']
            new_dict['include'] = q_d[key]['include']
            new_dict['diff'] = q_d[key]['diff']
            new_dict['year'] = q_d[key]['year']
            new_dict['norm'] = q_d[key]['norm']
            new_dict['data'] = added_data_dict[q_d[key]['Economic Indicators']]['Data']
            try:
                new_dict['data']['date'] = pd.to_datetime(new_dict['data']['date'], unit='D', origin='1899-12-30')
                new_dict['data']['date'] = new_dict['data']['date'].apply(lambda dt: dt.replace(day=1))
            except:
                pass
            var_dict[q_d[key]['Category']][q_d[key]['Economic Indicators']] = new_dict
        except:
            pass
    for key in m_d.keys():
        try:
            new_dict = {}
            new_dict['frequency'] = m_d[key]['Freq']
            new_dict['include'] = m_d[key]['include']
            new_dict['diff'] = m_d[key]['diff']
            new_dict['year'] = m_d[key]['year']
            new_dict['norm'] = m_d[key]['norm']
            new_dict['data'] = added_data_dict[m_d[key]['Economic Indicators']]['Data']
            try:
                new_dict['data']['date'] = pd.to_datetime(new_dict['data']['date'], unit='D', origin='1899-12-30')
                new_dict['data']['date'] = new_dict['data']['date'].apply(lambda dt: dt.replace(day=1))
            except:
                pass
            var_dict[m_d[key]['Category']][m_d[key]['Economic Indicators']] = new_dict
        except:
            pass
    return var_dict


# ### Data Processing

# In[19]:


#GDP column (target variable) processing

'''
- get data
- forecast missing values
- calculate filters, norm, sbbq
- get all stats
- splice data if multiple targets are included
- perform relevant processing steps
'''


def process_target_data(var_dict):
    
    target_dict = var_dict['target_variable']
    target_name = list(target_dict.keys())[0]
    df = target_dict[target_name]['data'].copy()

    if target_dict[target_name]['diff'] == True:
        if target_dict[target_name]['frequency'] == 'M':
            df[target_name] = df[target_name].pct_change(1)
        else:
            pass
    if target_dict[target_name]['year'] == True:
        if target_dict[target_name]['frequency'] == 'M':
            df[target_name] = df[target_name].pct_change(12)
        else:
            df[target_name] = df[target_name].pct_change(4)
        df = df.dropna()
        
    if target_dict[target_name]['norm'] == True:
        new_df = df[df.columns[1:]]
        normalized_df=(new_df-new_df.mean())/new_df.std()
        normalized_df.insert(0, 'date', df['date'])
        df = normalized_df.copy()
    
    original_target = list(df[target_name])
    #convert to gdp growth gap
    mean_g_growth = statistics.mean([x for x in list(df[target_name]) if str(x) != 'nan'])
    df['GDP_Gap'] = df[target_name].apply(lambda x: x - mean_g_growth)
    df[target_name] = original_target

    return df[['date','GDP_Gap']]


# In[20]:


#calculates all relevant filters for the target variable
def calculate_filters(df):
    gdp_ls = list(df['GDP_Gap'])
    new_gdp_ls = [x for x in gdp_ls if str(x) != 'nan']
    g_hp, _ = sm.tsa.filters.hpfilter(new_gdp_ls, 1600)
    g_cf, _ = sm.tsa.filters.cffilter(new_gdp_ls, 8, 32)
    g_ham_cycle, g_ham_trend = quantecon.hamilton_filter(new_gdp_ls, h=8, p=4)
    
    filter_dict = {'g_hp': [], 'g_cf': [], 'g_ham_cycle': [], 'g_ham_trend': []}
    filter_dict_init = {'g_hp': g_hp, 'g_cf': g_cf, 'g_ham_cycle': g_ham_cycle, 'g_ham_trend': g_ham_trend}
    count = 0
    for i in range(len(gdp_ls)):
        if math.isnan(gdp_ls[i]):
            for key in filter_dict.keys():
                filter_dict[key].append(np.nan)
        else:
            for key in filter_dict.keys():
                filter_dict[key].append(filter_dict_init[key][count])
            count += 1
    new_df = df.copy()        
    for key  in filter_dict.keys():
        new_df[key] = filter_dict[key]
    new_df['date'] = pd.PeriodIndex(new_df['date'], freq='Q')
    new_df = new_df.rename(columns={'date': 'quarter'})
    return new_df


# In[21]:


#calculate the normalized values of the GDP gap variable
def calculate_norm(df):
    col_ls = ['g_hp','g_cf','g_ham_cycle','GDP_Gap'] 
    for col in col_ls:
        d_ls = list(df[col])
        new_d_ls = [x for x in d_ls if str(x) != 'nan']
        d_mu = statistics.mean(new_d_ls)
        d_sd = statistics.stdev(new_d_ls)
        norm_ls = []
        for i in d_ls:
            norm_ls.append((i-d_mu)/d_sd)
        df[col+'_norm'] = norm_ls
    return df


# In[22]:


#calculate business cycle information using the TP detection pipeline
def calculate_sbbq(df):
    col_ls = ['g_hp_norm','g_cf_norm','g_ham_cycle_norm','GDP_Gap_norm'] 
    new_df = df[col_ls]
    new_df = cif.pipelineTPDetection(df = new_df.reset_index(), origColumns = col_ls, printDetails = False, showPlots = False, savePlots = None, saveLogs = None)
    for col in col_ls:
        df[col+'_point'] = list(new_df[col])
    return df


# In[23]:


#compute cycle statistics for GDP gap metric
def compute_total_cycles(point_ls):
    non_zero_ls = [x for x in point_ls if x != 0]
    count = 0
    for i in non_zero_ls:
        if i == 1:
            count+=1
    return count - 1

def compute_total_tp(point_ls):
    non_zero_ls = [x for x in point_ls if x != 0]
    non_zero_ls = [x for x in non_zero_ls if str(x) != 'nan']
    return len(non_zero_ls)

def compute_avg_slowdown_duration(point_ls):
    duration_ls = []
    for i in range(len(point_ls)):
        if point_ls[i] == 1:
            j = i
            count = 1
            while j < len(point_ls) and point_ls[j] != -1:
                count+=1
                j+=1
            duration_ls.append(count)
    try:
        mean_dur = statistics.mean(duration_ls)
    except:
        mean_dur = 0
    try:
        std_dur = statistics.stdev(duration_ls)
    except:
        std_dur = 0
    try:
        div = std_dur/mean_dur
    except:
        div = 0
    return mean_dur, div

def compute_avg_expansion_duration(point_ls):
    duration_ls = []
    for i in range(len(point_ls)):
        if point_ls[i] == -1:
            j = i
            count = 1
            while j < len(point_ls) and point_ls[j] != 1:
                count+=1
                j+=1
            duration_ls.append(count)
    try:
        mean_dur = statistics.mean(duration_ls)
    except:
        mean_dur = 0
    try:
        std_dur = statistics.stdev(duration_ls)
    except:
        std_dur = 0
    try:
        div = std_dur/mean_dur
    except:
        div = 0
    return mean_dur, div

def compute_avg_cycle_duration(s_duration, e_duration):
    return s_duration + e_duration

def compute_avg_slowdown_amplitude(point_ls, data_ls):
    amp_ls = []
    for i in range(len(point_ls)):
        if point_ls[i] == -1:
            amp_ls.append(data_ls[i])
    try:
        mean_amp = statistics.mean(amp_ls)
    except:
        mean_amp = 0
    try:
        std_amp = statistics.stdev(amp_ls)
    except:
        std_amp = 0
    try:
        div = std_amp/mean_amp
    except:
        div = 0
    return mean_amp, div

def compute_avg_expansion_amplitude(point_ls, data_ls):
    amp_ls = []
    for i in range(len(point_ls)):
        if point_ls[i] == 1:
            amp_ls.append(data_ls[i])
    try:
        mean_amp = statistics.mean(amp_ls)
    except:
        mean_amp = 0
    try:
        std_amp = statistics.stdev(amp_ls)
    except:
        std_amp = 0
    try:
        div = std_amp/mean_amp
    except:
        div = 0
    return mean_amp, div

def compute_speeds(e_dur, e_amp, s_dur, s_amp):
    try:
        expansion_speed = e_amp/e_dur
    except:
        expansion_speed = 0
    try:
        slowdown_speed = s_amp/s_dur
    except:
        slowdown_speed = 0
    return expansion_speed, slowdown_speed

def compute_last_quarter(df):
    new_df = df[df['GDP_Gap_norm_point'].notna()]
    qtr_ls = list(new_df['quarter'])
    point_ls = list(new_df['GDP_Gap_norm_point'])
    latest_qtr_data = qtr_ls[-1]
    point_ls.reverse()
    qtr_ls.reverse()
    last_qtr_peak = qtr_ls[point_ls.index(1)]
    num_qtrs_last_peak = point_ls.index(1)+1
    return {'Latest Quarter with Data': latest_qtr_data, 'Latest Quarter Peak': last_qtr_peak,
            'No. of Quarters from Last Peak': num_qtrs_last_peak}


# In[24]:


#get cycle statistics
def calculate_final_stats(df):
    final_dict = {}
    col_ls = ['g_hp_norm','g_cf_norm','g_ham_cycle_norm','GDP_Gap_norm']
    for col in col_ls:
        data_ls = list(df[col])
        point_ls = list(df[col+'_point'])
        total_cycles = compute_total_cycles(point_ls)
        total_tps = compute_total_tp(point_ls)
        avg_d_slowdown, cv_d_slowdown = compute_avg_slowdown_duration(point_ls)
        avg_d_expansion, cv_d_expansion = compute_avg_expansion_duration(point_ls)
        avg_d_cycles = compute_avg_cycle_duration(avg_d_slowdown, avg_d_expansion)
        avg_a_slowdown, cv_a_slowdown = compute_avg_slowdown_amplitude(point_ls, data_ls)
        avg_a_expansion, cv_a_expansion = compute_avg_expansion_amplitude(point_ls, data_ls)
        expansion_speed, slowdown_speed = compute_speeds(avg_d_expansion, avg_a_expansion, avg_d_slowdown, avg_a_slowdown)
        new_dict = {'Number of Cycles': total_cycles, 'Total No. of TPs': total_tps,
                    'Avg. Duration of Cycles': avg_d_cycles, 'Avg. Duration of Slowdown': avg_d_slowdown,
                    'Avg. Duration of Expansion': avg_d_expansion, 'Avg. Amplitude of Slowdown': avg_a_slowdown,
                    'Avg. Amplitude of Expansion': avg_a_expansion, 'Speed of Slowdown': slowdown_speed,
                    'Speed of Expansion': expansion_speed, 'CV of Slowdown Duration': cv_d_slowdown, 
                    'CV of Expansion Duration': cv_d_expansion, 'CV of Slowdown Amplitude': cv_a_slowdown,
                    'CV of Expansion Amplitude': cv_a_expansion}
        final_dict[col] = new_dict
        final_dict['summ_stat'] = compute_last_quarter(df)
    return final_dict


# In[25]:


#run the preceding cycle functions to get the updated normalized values, filters and cycle statistics for GDP growth gap
def get_cycles_information(df):
    with HiddenPrints():
        df = calculate_filters(df)
        df = calculate_norm(df)
        df = calculate_sbbq(df)
        stat_dict = calculate_final_stats(df)
    return df, stat_dict


# In[26]:


#Raw data processing. Processing steps are dependent on the users choices.

'''
- prune to selected indicators only
- forecast missing end values
- log/ln for selected indicators
- YoY or MoM differential for selected indicators
- normalize for selected indicators
- merge data with target variable
'''

def process_indicator_vars(var_dict):
    processed_data_dict = {
                 'consumption': [],
                 'exo': [],
                 'financial': [],
                 'government': [],
                 'investment': [],
                 'trade': []}
    
    contributions_dict = {
                 'consumption': [],
                 'exo': [],
                 'financial': [],
                 'government': [],
                 'investment': [],
                 'trade': []}
    
    for sector in var_dict.keys():
        if sector != 'target_variable':
            contributions_dict[sector] = list(var_dict[sector].keys())
            for indicator in var_dict[sector].keys():
                indicator_dict = var_dict[sector][indicator]
                df = indicator_dict['data'].copy()
                if indicator_dict['diff'] == True:
                    if indicator_dict['frequency'] == 'M':
                        df[indicator] = df[indicator].pct_change(1)
                    else:
                        pass
                if indicator_dict['year'] == True:
                    if indicator_dict['frequency'] == 'M':
                        df[indicator] = df[indicator].pct_change(12)
                    else:
                        df[indicator] = df[indicator].pct_change(4)
                    df = df.dropna()

                if indicator_dict['norm'] == True:
                    new_df = df[df.columns[1:]]
                    normalized_df=(new_df-new_df.mean())/new_df.std()
                    normalized_df.insert(0, 'date', df['date'])
                    df = normalized_df.copy()
                df.columns = ['date',indicator]
                processed_data_dict[sector].append(df)
    
    data_ls = []
    for key in processed_data_dict.keys():
        for i in processed_data_dict[key]:
            data_ls.append(i)
       
    final_df = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), data_ls)
    
    return final_df, contributions_dict


# In[27]:


#merge the target and indicator data series
def merge_data(added_data_dict, monthly_info, quarterly_info):
    var_dict = separate_indicator_target_data(added_data_dict, monthly_info, quarterly_info)
    df = process_target_data(var_dict)
    indicator_df, contributions_dict = process_indicator_vars(var_dict)
    cycles_df, stat_dict = get_cycles_information(df)
    
    cycles_df = cycles_df[['quarter','g_hp_norm','g_cf_norm','g_ham_cycle_norm','GDP_Gap']]
    cycles_df.columns = ['quarter','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']
    qtr_ls = list(cycles_df['quarter'])
    date_ls = []
    for q in qtr_ls:
        date_ls.append(q.to_timestamp())
    end_date = date_ls[-1]  + np.timedelta64(3, 'M')
    date_range = list(pd.date_range(date_ls[0], end_date, freq='MS'))
    new_dict = {'date': date_range}
    cycles_df['date'] = date_ls
    for col in ['date','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']:
        new_dict[col] = []
        col_ls = list(cycles_df[col])
        for i in range(len(date_ls)):
            new_dict[col] += [col_ls[i],col_ls[i],col_ls[i]]
    date_norm_df = pd.DataFrame(new_dict)
    while len(list(date_norm_df['date'])) < len(date_range):
        del date_range[-1]
    date_norm_df['date'] = date_range
    date_norm_df = date_norm_df[date_norm_df['date'] >= datetime(1990,1,1)]
    
    #merge EAI data
    if list(date_norm_df['date'])[-1] < list(indicator_df['date'])[-1]:
        df_ls = [indicator_df, date_norm_df]
    else:
        df_ls = [date_norm_df, indicator_df]
    data_monthly = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), df_ls)
    
    #create a quarterly version
    data_quarterly = data_monthly.copy()
    data_quarterly['quarter'] = data_quarterly['date'].dt.to_period('Q')
    data_quarterly = data_quarterly.groupby('quarter', as_index=False, sort=False).mean()
    data_quarterly = data_quarterly.drop(columns=['date'])
    
    return data_quarterly, data_monthly, stat_dict, contributions_dict


# ### Modelling

# In[28]:


#add a null value if the date is greater than start date
def add_nans(start_date, date, val):
    if date > start_date:
        return np.nan
    else:
        return val


# In[29]:


#forecast missing values using ARIMA
def forecast_missing_vals(df, col):
    mod = ARIMA((df[col].astype(float)), order=(1,0,0))
    res = mod.fit()
    return dict(res.predict(0,len(df[col])-1))


# In[30]:


#prepare all data for training
def prep_data(added_data_dict, monthly_info, quarterly_info):
    data_quarterly, data_monthly, cycles_dict, contributions_dict = merge_data(added_data_dict, monthly_info, quarterly_info)
    old_data_monthly = data_monthly.copy()
    old_data_quarterly = data_quarterly.copy()
    data_monthly = data_monthly.dropna(axis=1,how='all')
    data_monthly = data_monthly.loc[:, data_monthly.isnull().mean() < .85]
    earliest_actual_date = list(data_monthly[['date','GDP_Gap']].dropna()['date'])[0]
    latest_actual_date = list(data_monthly[['date','GDP_Gap']].dropna()['date'])[-1]
    latest_actual_quarter = list(data_monthly[['date','GDP_Gap']].dropna()['date'])[-3]
    data_monthly = data_monthly.drop(columns=['HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap'])
    gdp_df = data_quarterly[['quarter','GDP_Gap']].set_index('quarter').resample('M').interpolate(option='spline')
    gdp_df = gdp_df.to_timestamp()
    gdp_df = gdp_df.reset_index()
    gdp_df.columns = ['date','GDP_Gap']
    gdp_df = gdp_df[gdp_df['date'] <= latest_actual_date]
    gdp_df['GDP_Gap'] = gdp_df.apply(lambda x: add_nans(latest_actual_quarter, x['date'], x['GDP_Gap']),axis=1)
    missing_actual = list(forecast_missing_vals(gdp_df, 'GDP_Gap').values())
    actual_ls = list(gdp_df['GDP_Gap'])
    for i in range(len(actual_ls)):
        if math.isnan(actual_ls[i]):
            actual_ls[i] = missing_actual[i]
    gdp_df['GDP_Gap'] = actual_ls
    gdp_df = gdp_df[gdp_df['date'] >= earliest_actual_date]
    data_monthly = data_monthly.dropna()
    full_df = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), [data_monthly,gdp_df])
    #full_df = full_df[['date','GDP_Gap']+indicator_ls]
    new_earliest_date = list(data_monthly['date'])[0]
    full_df = full_df[full_df['date'] >= new_earliest_date]
    return full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict


# In[31]:


#use a given machine learning model to predict GDP growth gap monthly
def generate_ml_predictions(df, calcs_monthly, calcs_quarterly, stat_dict, contributions_dict, date_dict, model):
    #df, calcs_monthly, calcs_quarterly, stat_dict, contributions_dict = prep_data(added_data_dict, monthly_info, quarterly_info)
    model_df = df.copy().dropna()
    model_df = model_df.rename(columns = lambda x:re.sub('[^A-Za-z0-9_]+', '', x))
    preds = []
    if date_dict['start_date'] != '':
        model_df = model_df[model_df['date'] >= date_dict['start_date']].reset_index().drop(columns=['index'])
    if date_dict['end_date'] != '':
        if date_dict['end_date'][-2:] != '01':
            date_dict['end_date'] = date_dict['end_date'][:-2] +'01'
        num_oos =  len(model_df) - model_df[model_df['date'] == date_dict['end_date']].index[0]
    else:
        test_size = 0.5
        num_oos = round(len(model_df)*test_size)
    actual_vals = list(model_df['GDP_Gap'])
    window_preds = list(model_df['GDP_Gap'])
    shap_dict = {}
    count = 1
    model_df = model_df.set_index('date')
    for i in range(-num_oos, 0, 1):
        data_train = model_df[:i]
        data_test = model_df[i:]
        x_train = data_train.drop(columns=['GDP_Gap'])
        y_train = data_train[['GDP_Gap']]
        x_test = data_test.drop(columns=['GDP_Gap'])
        y_test = data_test[['GDP_Gap']]
        model.fit(x_train, y_train)
        try:
            pred = model.predict(x_test[:1])[0][0]
        except:
            pred = model.predict(x_test[:1])[0]
        if math.isnan(actual_vals[i]):
            actual_vals[i] = pred
            model_df['GDP_Gap'] = actual_vals
        window_preds[i] = pred
        explainer = shap.Explainer(model.predict, x_test[:1])
        shap_values = explainer(x_test[:1])
        shap_dict[count] = list(shap_values[0].values + shap_values[0].base_values/len(shap_values[0].values))
        count+=1
    model_df['EAI'] = window_preds
    shap_df = pd.DataFrame(shap_dict).T
    shap_df.columns = list(model_df.drop(columns = ['GDP_Gap','EAI']).columns)
    shap_df['date'] = list(model_df[-num_oos:].reset_index()['date'])
    combined_dict = {}
    for key in contributions_dict.keys():
        new_cols = []
        for col in contributions_dict[key]:
            new_cols.append(re.sub('[^A-Za-z0-9_]+', '', col))
        try:
            new_df = shap_df[new_cols]
        except:
            new_cols_ls = []
            for c in new_cols:
                if c in shap_df.columns:
                    new_cols_ls.append(c)
            new_df = shap_df[new_cols_ls]
        new_df[key] = new_df[list(new_df.columns)].sum(axis=1)
        combined_dict[key] = list(new_df[key])
    shap_df_2 = pd.DataFrame(combined_dict)
    shap_df_2['date'] = list(model_df[-num_oos:].reset_index()['date'])
    full_df = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), [model_df.reset_index(),shap_df_2])
    data_monthly = reduce(lambda left, right: pd.merge(left, right, on = ['date'], how = "outer"), [calcs_monthly,full_df[['date','EAI','consumption','exo','government','trade','financial','investment']]])
    return full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly


# In[32]:


def find_weights(overall, val):
    try:
        return val/overall
    except:
        return np.nan


# In[33]:


#get initial components per economic bucket using their shap values and weights
def get_eai_components(full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly):
    eai_components = data_monthly[['date','EAI','consumption','exo','government','trade','financial','investment']]
    rmse_df = data_monthly[['date','EAI','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']]
    ovr_weights = {}
    for col in ['consumption','exo','government','trade','financial','investment']:
        data_monthly[col+'_weight'] = data_monthly.apply(lambda x: find_weights(x['EAI'],x[col]),axis=1)
        ovr_weights[col] = data_monthly[col+'_weight'].mean()
    return eai_components, rmse_df, ovr_weights, calcs_quarterly, stat_dict


# In[34]:


#run the entire machine learning workflow
def ML_workflow(added_data_dict, date_dict, monthly_info, quarterly_info, model):
    print('data prep')
    full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict = prep_data(added_data_dict, monthly_info, quarterly_info)
    print('ML preds')
    full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly = generate_ml_predictions(full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict, date_dict, model)
    print('components calculations')
    eai_components, rmse_df, ovr_weights, calcs_quarterly, stat_dict = get_eai_components(full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly)
    print('done')
    return eai_components, rmse_df, ovr_weights, calcs_quarterly, stat_dict


# ### Components and Summary Calculations

# In[35]:


#calculate summary statistics for quarterly predictions
def calculate_quarterly_summaries(df):
    df['quarter'] = df['date'].dt.to_period('Q')
    quarterly_df = df.groupby('quarter', as_index=False, sort=False).mean()
    summary_dict = {}
    for col in ['HPBasedCycle','GDP_Gap','EAI']:
        summary_dict[col] = {}
        summary_dict[col]['mean'] = quarterly_df[col].mean()
        summary_dict[col]['stdev'] = quarterly_df[col].std()
    return quarterly_df, summary_dict


# In[36]:


#calculate overall components for EAI, GDP gap and HP based cycle metrics
def calculate_overall_components(df_monthly, summary_dict, ovr_weights_dict):
    eai_ls = list(df_monthly['EAI'])
    comps_dict = {'consumption': list(df_monthly['consumption']), 'exo': list(df_monthly['exo']), 'financial': list(df_monthly['financial']), 
                  'government': list(df_monthly['government']), 'investment': list(df_monthly['investment']), 'trade': list(df_monthly['trade'])}
    for comp in ['GDP_Gap','HPBasedCycle']:
        new_eai_ls = []
        for i in range(len(eai_ls)):
            new_eai_ls.append(summary_dict[comp]['mean']+(eai_ls[i]-summary_dict['EAI']['mean'])*(summary_dict[comp]['stdev']/summary_dict['EAI']['stdev']))
        df_monthly['EAI_'+comp] = new_eai_ls
        for key in comps_dict.keys():
            comp_ls = comps_dict[key]
            new_comp_ls = []
            for i in range(len(comp_ls)):
                new_comp_ls.append(new_eai_ls[i]*ovr_weights_dict[key])
            df_monthly[key+'_'+comp] = new_comp_ls
    return df_monthly


# In[37]:


#run all previous functions to generate the overall components dataframe
def get_overall_components(added_data_dict, date_dict, monthly_info, quarterly_info, model, verbose = False):
    if verbose == True:
        print('Extracting data and calculating EAI values for ' + str(country) + '...')
    eai_comps, rmse_df, ovr_weights, calcs_quarterly, cycle_stat_dict = ML_workflow(added_data_dict, date_dict, monthly_info, quarterly_info, model)
    if verbose == True:
        print('Calculating quarterly summaries of real data for ' + str(country) + '...')
    new_rmse_df, summary_dict = calculate_quarterly_summaries(rmse_df)
    if verbose == True:
        print('Calculating overall component figures for ' + str(country) + '...')
    overall_eai_components = calculate_overall_components(eai_comps, summary_dict, ovr_weights)
    quarterly_data = calcs_quarterly[['quarter','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']]
    return overall_eai_components, quarterly_data, cycle_stat_dict, summary_dict


# ### Data Forecasting

# In[38]:


#match the lengths of a dictionary using null values to prepare for dataframe conversion
def match_dict_lens(dict_to_match):
    ls_lens = []
    for key in dict_to_match.keys():
        ls_lens.append(len(dict_to_match[key]))
    max_len = max(ls_lens)
    for key in dict_to_match.keys():
        if len(dict_to_match[key]) < max_len:
            while len(dict_to_match[key]) < max_len:
                dict_to_match[key].append(np.nan)
    return dict_to_match


# In[39]:


#predict the next x months using a linear forecasting method (ARIMA)
def predict_EAI(df, months_to_predict):
    df = df.reset_index()[['date','EAI']]
    mod = ARIMA((df['EAI'].astype(float)), order=(1,0,0))
    res = mod.fit()
    max_rows = len(df)-1
    preds = list(res.predict(max_rows, max_rows + months_to_predict))[-months_to_predict:]
    ci95 = res.get_forecast(months_to_predict).conf_int(alpha=0.05)[:months_to_predict]
    ci99 = res.get_forecast(months_to_predict).conf_int(alpha=0.01)[:months_to_predict]
    eai_ls = list(df['EAI'])
    ci95_lower_ls = eai_ls + list(ci95['lower EAI'])
    ci95_upper_ls = eai_ls + list(ci95['upper EAI'])
    ci99_lower_ls = eai_ls + list(ci99['lower EAI'])
    ci99_upper_ls = eai_ls + list(ci99['upper EAI'])
    pred_ls = eai_ls + preds
    date_ls = list(df['date'])
    end_date = date_ls[-1]  + np.timedelta64(months_to_predict+1, 'M')
    date_range = list(pd.date_range(date_ls[0], end_date, freq='MS'))
    new_df_dict = {'date': date_range, 'EAI': eai_ls, 'pred': pred_ls, 
                   '95%CILowerLimit': ci95_lower_ls, '95%CIUpperLimit': ci95_upper_ls,
                   '99%CILowerLimit': ci99_lower_ls, '99%CIUpperLimit': ci99_upper_ls}
    new_df_dict = match_dict_lens(new_df_dict)
    return pd.DataFrame(new_df_dict)


# In[40]:


#Run previous functions to get the components, statistics, and predictions
def get_EAI_update(added_data_dict, date_dict, monthly_info, quarterly_info, model, months_to_predict = 3, verbose = False):
    eai_components, real_quarterly_data, cycle_stats, summary_dict = get_overall_components(added_data_dict, date_dict, monthly_info, quarterly_info, model, verbose=verbose)
    try:
        eai_components['date'] = pd.to_datetime(eai_components['date'])
        eai_components = eai_components.sort_values(by=['date'], ascending = True)
    except:
        pass
    new_components = eai_components.copy()
    if verbose == True:
        print('Generating EAI predictions and confidence intervals for the next ' + str(months_to_predict) + ' months for ' + str(country) + '...')
    eai_preds = predict_EAI(new_components, months_to_predict)
    real_quarterly_data = real_quarterly_data.set_index('quarter').dropna(axis = 0, how = 'all').reset_index()
    if verbose == True:
        print('Finished generating figures for ' + str(country) + '.')
    return eai_components, real_quarterly_data, cycle_stats, eai_preds, summary_dict


# ### Additional Calculations

# In[41]:


#write a dataframe into an excel file and sheet
def write_excel(filename, sheet_name, df):
    if not os.path.isfile(filename):
        wb = Workbook()
        ws = wb.active
        wb.save(filename)
        wb.close()
    og_cols = list(df.columns)
    cols=pd.Series(df.columns)
    for dup in cols[cols.duplicated()].unique(): 
        cols[cols[cols == dup].index.values.tolist()] = [dup + '.' + str(i) if i != 0 else dup for i in range(sum(cols == dup))]
    df.columns=cols
    wb = load_workbook(filename)
    if sheet_name in wb.get_sheet_names():
        sheet_to_del=wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet_to_del)
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    w_exc_dict = {}
    cols = list(df.columns)
    for c in cols:
        w_exc_dict[c] = list(df[c])
    col_count = 1
    for d in w_exc_dict.keys():
        ws.cell(row=1, column=col_count).value = og_cols[col_count-1]
        for i in range(len(w_exc_dict[d])):
            try:
                w_exc_dict[d][i] = w_exc_dict[d][i].replace(tzinfo=None)
            except:
                pass
            ws.cell(row=i+2, column=col_count).value = w_exc_dict[d][i]
        try:
            m_row = ws.max_row
            for i in range(m_row+1):
                try:
                    ws.cell(row=i+1, column=col_count).value = ws.cell(row=i+1, column=col_count).value.strftime("%m/%d/%Y")
                except:
                    pass
        except:
            pass
        col_count+=1
    wb.save(filename)
    wb.close()


# In[42]:


#prepare the average cycles, eai components, eai predictions, quarterly eai and quarterly gdp gap sheets
def prepare_country_file(components, qtr_data, stats, preds, summ_dict, filename):
    #Average Cycle
    metrics = ['Total No. of TPs','Ave. duration of cycles (qtr)','Ave. duration of slowdown (qtr)','Ave. duration of expansion (qtr)','Ave. amplitude of slowdown','Ave. amplitude of expansion','Speed of slowdown','Speed of expansion','CV_d of slowdown','CV_d of expansion','CV_a of slowdown','CV_a of expansion','Ave. RGDP growth of slowdown','Ave. RGDP growth of expansion','Ave. credit/gdp growth of slowdown','Ave. credit/gdp growth of expansion','Number of quarters from last peak']
    values = list(stats['GDP_Gap_norm'].values())[1:] + [np.nan,np.nan,np.nan,np.nan,stats['summ_stat']['No. of Quarters from Last Peak']]
    avg_cycles_df = pd.DataFrame({'Metric': metrics, 'Value': values})
    
    #EAI_Components
    comps = components.copy()
    comps.columns = ['month','EAI','CONSUMPTION','EXO_INTERNATIONAL','GOVERNMENT','TRADE','FINANCIAL','INVESTMENT',
                          'EAI_GDP_gap','CONSUMPTION_GDP_gap','EXO_INTERNATIONAL_GDP_gap','FINANCIAL_GDP_gap','GOVERNMENT_GDP_gap','INVESTMENT_GDP_gap','TRADE_GDP_gap',
                          'EAI_HP','CONSUMPTION_HP','EXO_INTERNATIONAL_HP','FINANCIAL_HP','GOVERNMENT_HP','INVESTMENT_HP','TRADE_HP']
    comps = comps[['month','EAI','CONSUMPTION','EXO_INTERNATIONAL','FINANCIAL','GOVERNMENT','INVESTMENT','TRADE',
                          'EAI_GDP_gap','CONSUMPTION_GDP_gap','EXO_INTERNATIONAL_GDP_gap','FINANCIAL_GDP_gap','GOVERNMENT_GDP_gap','INVESTMENT_GDP_gap','TRADE_GDP_gap',
                          'EAI_HP','CONSUMPTION_HP','EXO_INTERNATIONAL_HP','FINANCIAL_HP','GOVERNMENT_HP','INVESTMENT_HP','TRADE_HP']]
    comps_dict = {}
    for col in list(comps.columns):
        comps_dict[col] = list(comps[col])

    comps_dict['blank1'] = [np.nan]
    comps_dict['blank2'] = [np.nan]
    comps_dict['Variable'] = ['GDP_Gap','HPBasedCycle','WALS']
    comps_dict['Mean'] = [summ_dict['GDP_Gap']['mean'],summ_dict['HPBasedCycle']['mean'],summ_dict['EAI']['mean']]
    comps_dict['Std. Dev.'] = [summ_dict['GDP_Gap']['stdev'],summ_dict['HPBasedCycle']['stdev'],summ_dict['EAI']['stdev']]
    comps_dict = match_dict_lens(comps_dict)
    eai_components_df = pd.DataFrame(comps_dict)

    
    #if column contains only 0, then change all values to NaN
    for col in list(eai_components_df.columns):
        if (eai_components_df[col].dropna() == 0).all():
            eai_components_df[col] = np.nan

    #EAI_prediction
    new_preds = preds.copy()
    new_preds.columns = ['month','EAI','PredictedEAI','95%CiLowerLimit','95%CiUpperLimit','99%CiLowerLimit', '99%CiUpperLimit']
    
    #EAI_quarterly
    eai_qtr = eai_components_df[['month','EAI']]
    eai_qtr['qtr'] = eai_qtr['month'].dt.to_period('Q')
    eai_qtr = eai_qtr.groupby('qtr', as_index=False, sort=False).mean()
    eai_qtr = eai_qtr.drop(columns=['month'])
    eai_qtr = eai_qtr[['qtr','EAI']]
    eai_qtr['qtr'] = pd.PeriodIndex(eai_qtr['qtr'], freq='Q').to_timestamp()
    
    #GDP_gap_quarterly
    q_df = qtr_data.copy()
    q_df['qtr'] = pd.PeriodIndex(q_df['quarter'], freq='Q').to_timestamp()
    q_df = q_df[q_df['qtr'] >= datetime(2000,1,1)]
    q_df_gdp = q_df[['qtr','GDP_Gap','HPBasedCycle']]
    
    #filename = "C:/Users/Patrick Jaime Simba/Desktop/ADB Files/Notebooks/ML Time Series/ML Methodology/Updated Files V2/EAI_" + country + ".xlsx"
    write_excel(filename, 'Average Cycles', avg_cycles_df)
    write_excel(filename, 'EAI_Components', eai_components_df)
    write_excel(filename, 'EAI_prediction', new_preds)
    write_excel(filename, 'EAI_quarterly', eai_qtr)
    write_excel(filename, 'GDP_gap_quarterly', q_df_gdp)
    
    return avg_cycles_df, eai_components_df, new_preds, eai_qtr, q_df_gdp


# In[43]:


#use previous functions to prepare all relevant data and save files
def generate_EAI_update(added_data_dict, date_dict, monthly_info, quarterly_info, model, filename, verbose = False):
    components, qtr_data, stats, preds, summary_dict = get_EAI_update(added_data_dict, date_dict, monthly_info, quarterly_info, model)
    if verbose:
        print("Preparing country EAI excel file...")
    avg_cycles, eai_comps, eai_preds, eai_qtrly, gdp_hp_qtrly = prepare_country_file(components, qtr_data, stats, preds, summary_dict, filename)
    if verbose:
        print("EAI_"+country+".xlsx file successfully generated.")


# ### Additional Excel Prep

# In[44]:


#open an excel file and particular sheet
def open_sheet(filename, sheet):
    df = pd.read_excel(filename, sheet_name=sheet)
    return df


# In[45]:


#find the coordinates for each predicted value and GDP gap for the business cycle dials
def find_eai_dial_coordinates(curr, prev, axis, trend):
    try:
        if axis == 'x':
            if (curr >= prev) and (curr >= trend):
                return 1.7
            elif (curr >= prev) and (curr < trend):
                return 0.3
            elif (curr < prev) and (curr >= trend):
                return 2.5
            elif (curr < prev) and (curr < trend):
                return -0.5
            else:
                return ''
        elif axis == 'y':
            if (curr >= prev) and (curr >= trend):
                return 2.25
            elif (curr >= prev) and (curr < trend):
                return 2.25
            elif (curr < prev) and (curr >= trend):
                return 1.35
            elif (curr < prev) and (curr < trend):
                return 1.35
            else:
                return ''
    except:
        return ''


# In[46]:


#find the coordinates for the sector component dials
def find_component_dial_coordinates(curr, axis):
    try:
        if axis == 'x':
            if (curr > 0):
                return 2.2
            elif (curr < 0):
                return -0.2
            elif (curr == 0):
                return 1
            else:
                return ''
        elif axis == 'y':
            if (curr > 0):
                return 2.05
            elif (curr < 0):
                return 2.05
            elif (curr == 0):
                return 2.5
            else:
                return ''
    except:
        return ''


# In[47]:


#calculate and create excel sheets for the final file
def create_dial_coordinates_sheet(filename, trend=0):
    components = open_sheet(filename, 'EAI_Components')
    #components = components.dropna(subset=['EAI'])
    comps_dict = {'eai': list(components['EAI']), 'con': list(components['CONSUMPTION']), 
                 'exo': list(components['EXO_INTERNATIONAL']), 'fin': list(components['FINANCIAL']),
                 'gov': list(components['GOVERNMENT']), 'inv': list(components['INVESTMENT']), 
                 'tra': list(components['TRADE'])}
    new_data_dict = {'month': list(components['month'])}
    for key in comps_dict.keys():
        new_data_dict[key] = comps_dict[key]
        prev_ls = ['']
        for i in range(len(comps_dict[key])-1):
            prev_ls.append(comps_dict[key][i])
        new_data_dict['prev_'+key] = prev_ls
        x_ls = []
        y_ls = []
        for i in range(len(new_data_dict[key])):
            if key == 'eai':
                x_ls.append(find_eai_dial_coordinates(new_data_dict[key][i], new_data_dict['prev_'+key][i], 'x', trend))
                y_ls.append(find_eai_dial_coordinates(new_data_dict[key][i], new_data_dict['prev_'+key][i], 'y', trend))
            else:
                x_ls.append(find_component_dial_coordinates(new_data_dict[key][i], 'x'))
                y_ls.append(find_component_dial_coordinates(new_data_dict[key][i], 'y'))
        new_data_dict['x_param_'+key] = x_ls
        new_data_dict['y_param_'+key] = y_ls
    return pd.DataFrame(new_data_dict)


# In[48]:


def create_pred_vals_sheet(filename):
    predictions = open_sheet(filename, 'EAI_prediction')
    gdp_gap = open_sheet(filename, 'GDP_gap_quarterly')
    gdp_gap = gdp_gap[['qtr','GDP_Gap']]
    gdp_gap.columns = ['month','GDP_Gap']
    gdp_gap['month'] = pd.to_datetime(gdp_gap['month'])
    pred_dict = {'month': list(predictions['month']), 'EAI': list(predictions['EAI']), 
                 'PredictedEAI': list(predictions['EAI'])[:-3]+list(predictions['PredictedEAI'])[-3:]}
    for i in range(len(pred_dict['month'])):
        if type(pred_dict['month'][i]) == int:
            pred_dict['month'][i] = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + pred_dict['month'][i] - 2)
    
    ucl95 = []
    lcl95 = []
    ucl99 = []
    lcl99 = []
    for i in range(len(pred_dict['EAI'][:-4])):
        ucl95.append(np.nan)
        lcl95.append(np.nan)
        ucl99.append(np.nan)
        lcl99.append(np.nan)
    ucl95.append(list(predictions['EAI'])[-4])
    lcl95.append(list(predictions['EAI'])[-4])
    ucl99.append(list(predictions['EAI'])[-4])
    lcl99.append(list(predictions['EAI'])[-4])

    for j in [-3,-2,-1]:
        ucl95.append(list(predictions['95%CiUpperLimit'])[j])
        lcl95.append(list(predictions['95%CiLowerLimit'])[j])
        ucl99.append(list(predictions['99%CiUpperLimit'])[j])
        lcl99.append(list(predictions['99%CiLowerLimit'])[j])

    
    pred_dict['UCL95'] = ucl95
    pred_dict['LCL95'] = lcl95
    pred_dict['UCL99'] = ucl99
    pred_dict['LCL99'] = lcl99
    
    new_df = pd.DataFrame(pred_dict)
    new_df['month'] = pd.to_datetime(new_df['month'])
    final_df = reduce(lambda left, right: pd.merge(left, right, on = ['month'], how = "outer"), [new_df, gdp_gap])
    #final_df['month'] = pd.to_datetime(final_df['month'])
    #final_df = final_df.sort_values(by=['month'], ascending = True)
    return final_df


# In[49]:


def create_quarterly_computations_sheet(filename):
    eai_df = open_sheet(filename, 'EAI_Components')
    eai_df_2 = open_sheet(filename, 'EAI_quarterly')
    gdp_df = open_sheet(filename, 'GDP_gap_quarterly')
    full_df_1 = reduce(lambda left, right: pd.merge(left , right,on = ["qtr"],how = "outer"),[eai_df_2,gdp_df])
    d_dict = {'month': list(eai_df['month']), 'EAI': list(eai_df['EAI']), 'quarter': list(full_df_1['qtr']), 
              'EAI_quarterly': list(full_df_1['EAI']), 'GDP_gap_quarterly': list(full_df_1['GDP_Gap']),
              'HPBasedCycle_quarterly': list(full_df_1['HPBasedCycle'])}
    d_dict = match_dict_lens(d_dict)
    return pd.DataFrame(d_dict)


# In[50]:


def create_quarterly_data_sheet(filename, trend = 0):
    quarterly_comps = open_sheet(filename, 'Quarterly_computations')
    eai_quarterly = open_sheet(filename, 'EAI_quarterly')
    qtr_dict = {}
    qtr_dict['quarter'] = list(eai_quarterly['qtr'])
    
    eai_ls = list(quarterly_comps['EAI_quarterly'])
    #eai_ls = [item for item in eai_ls if not(math.isnan(item))==True]
    qtr_dict['EAI_quarterly'] = eai_ls
    prev_eai_ls = ['']
    for i in eai_ls[:-1]:
        prev_eai_ls.append(i)
    qtr_dict['prev_eai_quarterly'] = prev_eai_ls
    qtr_dict['x_param_eai_quarterly'] = []
    qtr_dict['y_param_eai_quarterly'] = []
    for i in range(len(qtr_dict['EAI_quarterly'])):
        qtr_dict['x_param_eai_quarterly'].append(find_eai_dial_coordinates(qtr_dict['EAI_quarterly'][i], qtr_dict['prev_eai_quarterly'][i], 'x', trend))
        qtr_dict['y_param_eai_quarterly'].append(find_eai_dial_coordinates(qtr_dict['EAI_quarterly'][i], qtr_dict['prev_eai_quarterly'][i], 'y', trend))
        
    gdp_ls = list(quarterly_comps['GDP_gap_quarterly'])
    #gdp_ls = [item for item in gdp_ls if not(math.isnan(item))==True]
    while math.isnan(gdp_ls[-1]):
            del gdp_ls[-1]
    qtr_dict['GDP_gap_quarterly'] = gdp_ls
    prev_gdp_ls = ['']
    for i in gdp_ls[:-1]:
        prev_gdp_ls.append(i)
    qtr_dict['prev_gdp_gap_quarterly'] = prev_gdp_ls
    qtr_dict['x_param_gdp_gap_quarterly'] = []
    qtr_dict['y_param_gdp_gap_quarterly'] = []
    for i in range(len(qtr_dict['GDP_gap_quarterly'])):
        qtr_dict['x_param_gdp_gap_quarterly'].append(find_eai_dial_coordinates(qtr_dict['GDP_gap_quarterly'][i], qtr_dict['prev_gdp_gap_quarterly'][i], 'x', trend))
        qtr_dict['y_param_gdp_gap_quarterly'].append(find_eai_dial_coordinates(qtr_dict['GDP_gap_quarterly'][i], qtr_dict['prev_gdp_gap_quarterly'][i], 'y', trend))
    
    try:
        hp_ls = list(quarterly_comps['HPBasedCycle_quarterly'])
        #hp_ls = [item for item in hp_ls if not(math.isnan(item))==True]
        while math.isnan(hp_ls[-1]):
            del hp_ls[-1]
        qtr_dict['HPBasedCycle_quarterly'] = hp_ls
        prev_hp_ls = ['']
        for i in hp_ls[:-1]:
            prev_hp_ls.append(i)
        qtr_dict['prev_HPBasedCycle_quarterly'] = prev_hp_ls
        qtr_dict['x_param_HPBasedCycle_quarterly'] = []
        qtr_dict['y_param_HPBasedCycle_quarterly'] = []
        for i in range(len(qtr_dict['HPBasedCycle_quarterly'])):
            qtr_dict['x_param_HPBasedCycle_quarterly'].append(find_eai_dial_coordinates(qtr_dict['HPBasedCycle_quarterly'][i], qtr_dict['prev_HPBasedCycle_quarterly'][i], 'x', trend))
            qtr_dict['y_param_HPBasedCycle_quarterly'].append(find_eai_dial_coordinates(qtr_dict['HPBasedCycle_quarterly'][i], qtr_dict['prev_HPBasedCycle_quarterly'][i], 'y', trend))
    except:
        pass
    
    for key in qtr_dict.keys():
        #print(key)
        while pd.isnull(qtr_dict[key][-1]):
            del qtr_dict[key][-1]
    
    qtr_dict = match_dict_lens(qtr_dict)
    return pd.DataFrame(qtr_dict)


# In[51]:


def segment1_formula(val, amp, x_intercept, shift):
    b_val = x_intercept/math.acos(0)
    return amp * math.cos(val/b_val)

def segment2_formula(val, amp, x_intercept, shift):
    b_val = x_intercept/math.acos(0)
    return -amp * math.cos(val/b_val)

def segment3_formula(val, amp, x_intercept, shift):
    b_val = x_intercept/math.acos(0)
    return amp * math.cos((val-shift)/b_val)

def segment4_formula(val, amp, x_intercept, shift):
    b_val = x_intercept/math.acos(0)
    return -amp * math.cos((val-shift)/b_val)


# In[52]:


def find_latest_quarter_constant(latest, prev, middle, maximum):
    if latest > prev and latest > 0:
        return maximum-((maximum-middle)/4)
    elif latest > prev and latest < 0:
        return middle+((maximum-middle)/4)
    elif latest < prev and latest < 0:
        return middle-((maximum-middle)/4)
    elif latest < prev and latest > 0:
        return ((maximum-middle)/4)
    else:
        return 0


# In[53]:


def check_cycle_position_v0(curr, prev, trend):
    if curr > trend and curr >= prev:
        return 12.8
    elif curr > trend and curr < prev:
        return 2.2
    elif curr < trend and curr >= prev:
        return 9.0
    else:
        return 5.9
    
def check_cycle_position(curr, prev, trend):
    if curr > trend and curr >= prev:
        return 9.0
    elif curr > trend and curr < prev:
        return 12.8
    elif curr < trend and curr >= prev:
        return 5.9
    else:
        return 2.2


# In[54]:


def create_curved_data_sheet(filename, trend = 0):
    quarterly_data_df = open_sheet(filename, 'Quarterly_data')
    avg_cycles = open_sheet(filename, 'Average Cycles')
    cycles_dict = {}
    for i in range(len(list(avg_cycles['Metric']))):
        cycles_dict[list(avg_cycles['Metric'])[i]] = list(avg_cycles['Value'])[i]
    latest_quarterly_eai = [x for x in list(quarterly_data_df['EAI_quarterly']) if str(x) != 'nan'][-1]
    previous_quarterly_eai = [x for x in list(quarterly_data_df['EAI_quarterly']) if str(x) != 'nan'][-2]
    latest_quarter_val_ls = [x for x in list(quarterly_data_df['quarter']) if str(x) != 'nan']
    gdp_gap_quarterly_ls = [x for x in list(quarterly_data_df['GDP_gap_quarterly']) if str(x) != 'nan']
    latest_quarterly_gdp = gdp_gap_quarterly_ls[-1]
    previous_quarterly_gdp = gdp_gap_quarterly_ls[-2]
    curved_dict = {'x_initial': [], 'y_initial': [], 'x': [], 'y': [], 'latest_y_val': [],
                 'latest_quarterly_month': [latest_quarter_val_ls[-1]], 'latest_quarterly_month_gdp': [latest_quarter_val_ls[-2]], 'latest_quarterly_eai': [latest_quarterly_eai], 
                 'previous_quarterly_eai': [previous_quarterly_eai], 'latest_quarterly_gdp': [latest_quarterly_gdp], 
                 'previous_quarterly_gdp': [previous_quarterly_gdp], 'latest_quarterly_quarter': [ round(find_latest_quarter_constant(latest_quarterly_eai, previous_quarterly_eai, cycles_dict['Ave. duration of slowdown (qtr)'], cycles_dict['Ave. duration of cycles (qtr)']),1)]}
    curved_dict['x_initial'] = [0, cycles_dict['Ave. duration of slowdown (qtr)'], cycles_dict['Ave. duration of cycles (qtr)']]
    curved_dict['y_initial'] = [cycles_dict['Ave. amplitude of expansion'], -cycles_dict['Ave. amplitude of slowdown'], cycles_dict['Ave. amplitude of expansion']]
    x_ls = list(np.round(np.arange(0.0, round(15,1)+0.1, 0.1),1))
    y_ls = []
    for i in x_ls:
        y_ls.append(3*math.cos(0.5*(i) - 5.313))
    curved_dict['x'] = x_ls
    curved_dict['y'] = y_ls
    cycle_pos = check_cycle_position(curved_dict['latest_quarterly_eai'][0], curved_dict['previous_quarterly_eai'][0], trend)
    #print(cycle_pos)
    latest_y_val_ls = []
    for i in range(len(x_ls)):
        if x_ls[i] == cycle_pos:
            latest_y_val_ls.append(y_ls[i])
        else:
            latest_y_val_ls.append(np.nan)
    curved_dict['latest_y_val'] = latest_y_val_ls
    latest_y_val_ls_gdp = []
    cycle_pos_gdp = check_cycle_position(curved_dict['latest_quarterly_gdp'][0], curved_dict['previous_quarterly_gdp'][0], trend)
    for i in range(len(x_ls)):
        if x_ls[i] == cycle_pos_gdp:
            latest_y_val_ls_gdp.append(y_ls[i])
        else:
            latest_y_val_ls_gdp.append(np.nan)
    curved_dict['latest_y_val_gdp'] = latest_y_val_ls_gdp
    curved_dict = match_dict_lens(curved_dict)
    curved_df = pd.DataFrame(curved_dict)
    return curved_df


# In[55]:


def create_gdp_curved_data_sheet(filename, trend = 0):
    quarterly_data_df = open_sheet(filename, 'Quarterly_data')
    avg_cycles = open_sheet(filename, 'Average Cycles')
    cycles_dict = {}
    for i in range(len(list(avg_cycles['Metric']))):
        cycles_dict[list(avg_cycles['Metric'])[i]] = list(avg_cycles['Value'])[i]
    gdp_gap_quarterly_ls = [x for x in list(quarterly_data_df['GDP_gap_quarterly']) if str(x) != 'nan']
    latest_quarterly_eai = gdp_gap_quarterly_ls[-1]
    previous_quarterly_eai = gdp_gap_quarterly_ls[-2]
    latest_quarter_val_ls = [x for x in list(quarterly_data_df['quarter']) if str(x) != 'nan']
    curved_dict = {'x_initial': [], 'y_initial': [], 'x': [], 'y': [], 'latest_y_val': [],
                 'latest_quarterly_month': [latest_quarter_val_ls[-1]], 'latest_quarterly_gdp': [latest_quarterly_eai], 
                 'previous_quarterly_gdp': [previous_quarterly_eai], 'latest_quarterly_quarter': [ round(find_latest_quarter_constant(latest_quarterly_eai, previous_quarterly_eai, cycles_dict['Ave. duration of slowdown (qtr)'], cycles_dict['Ave. duration of cycles (qtr)']),1)]}
    curved_dict['x_initial'] = [0, cycles_dict['Ave. duration of slowdown (qtr)'], cycles_dict['Ave. duration of cycles (qtr)']]
    curved_dict['y_initial'] = [cycles_dict['Ave. amplitude of expansion'], -cycles_dict['Ave. amplitude of slowdown'], cycles_dict['Ave. amplitude of expansion']]
    x_ls = list(np.round(np.arange(0.0, round(15,1)+0.1, 0.1),1))
    y_ls = []
    for i in x_ls:
        y_ls.append(3*math.cos(0.5*(i) - 5.313))
    curved_dict['x'] = x_ls
    curved_dict['y'] = y_ls
    cycle_pos = check_cycle_position(curved_dict['latest_quarterly_gdp'][0], curved_dict['previous_quarterly_gdp'][0], trend)
    latest_y_val_ls = []
    #print(curved_dict)
    for i in range(len(x_ls)):
        if x_ls[i] == cycle_pos:
            latest_y_val_ls.append(y_ls[i])
        else:
            latest_y_val_ls.append(np.nan)
    curved_dict['latest_y_val'] = latest_y_val_ls
    curved_dict = match_dict_lens(curved_dict)
    curved_df = pd.DataFrame(curved_dict)
    return curved_df


# In[56]:


def create_cycle_bands_sheet(filename):
    cycle_data = open_sheet(filename, 'curved_data')
    cycle_dict = {
                 'blue_band_min': [list(cycle_data['x_initial'])[0]],
                 'blue_band_max': [list(cycle_data['x_initial'])[1]/2],
                 'red_band_min': [list(cycle_data['x_initial'])[1]/2],
                 'red_band_max': [list(cycle_data['x_initial'])[1]],
                 'yellow_band_min': [list(cycle_data['x_initial'])[1]],
                 'yellow_band_max': [(list(cycle_data['x_initial'])[2]/4) + list(cycle_data['x_initial'])[1]],
                 'green_band_min': [(list(cycle_data['x_initial'])[2]/4) + list(cycle_data['x_initial'])[1]],
                 'green_band_max': [list(cycle_data['x_initial'])[2]]}
    return pd.DataFrame(cycle_dict)


# In[57]:


#run previous functions to prepare additional calculation sheets for the dashboard
def prepare_sheets(filepath):
    
    trend = 0
    
    dial_sheet = create_dial_coordinates_sheet(filepath, trend)
    write_excel(filepath, 'Dial_Coordinates', dial_sheet)
    
    pred_sheet = create_pred_vals_sheet(filepath)
    pred_sheet['gdp_trend'] = trend
    write_excel(filepath, 'pred_vals', pred_sheet)
    
    qtr_comps_sheet = create_quarterly_computations_sheet(filepath)
    write_excel(filepath, 'Quarterly_computations', qtr_comps_sheet)
    
    qtr_data_sheet = create_quarterly_data_sheet(filepath, trend)
    write_excel(filepath, 'Quarterly_data', qtr_data_sheet)
    
    curved_sheet = create_curved_data_sheet(filepath, trend)
    write_excel(filepath, 'curved_data', curved_sheet)
    
    gdp_curved_sheet = create_gdp_curved_data_sheet(filepath, trend)
    write_excel(filepath, 'curved_gdp_data', curved_sheet)
    
    band_sheet = create_cycle_bands_sheet(filepath)
    write_excel(filepath, 'cycle_bands', band_sheet)


# ### Final Functions

# In[58]:


#dictionary containing all possible model types
full_model_dictionary = {
     'Linear Regression': LinearRegression(),
     'Random Forest': RandomForestRegressor(),
     'Light Gradient Boosting Method (LGBM)': LGBMRegressor(),
     'Extreme Boosting (XGBoost)': XGBRegressor(),
     'CatBoost': CatBoostRegressor(),
     'Stochastic Gradient Descent (SGDR)': SGDRegressor(),
     'Kernel Ridge': KernelRidge(),
     'Elastic Net': ElasticNet(),
     'Bayesian Ridge': BayesianRidge(),
     'Gradient Boosting': GradientBoostingRegressor(),
     'Support Vector Machine (SVM)': SVR()
}


# In[59]:


#run all back end operations
def operations_func(added_data_dict, date_dict, monthly_info, quarterly_info, model_name, filepath, log_str):
    
    model = full_model_dictionary[model_name]
    
    log_str += 'Preparing data for ' + model_name +' model.\n'
    full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict = prep_data(added_data_dict, monthly_info, quarterly_info)
    
    log_str += 'Generating predictions.\n'
    full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly = generate_ml_predictions(full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict, date_dict, model)
    
    log_str += 'Calculating initial components.\n'
    eai_comps, rmse_df, ovr_weights, calcs_quarterly, cycle_stats = get_eai_components(full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly)
    
    log_str += 'Calculating data summaries.\n'
    new_rmse_df, summary_dict = calculate_quarterly_summaries(rmse_df)
    
    log_str += 'Calculating overall components.\n'
    eai_components = calculate_overall_components(eai_comps, summary_dict, ovr_weights)
    real_quarterly_data = calcs_quarterly[['quarter','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']]
    
    try:
        eai_components['date'] = pd.to_datetime(eai_components['date'])
        eai_components = eai_components.sort_values(by=['date'], ascending = True)
    except:
        pass
    new_components = eai_components.copy()
    
    log_str += 'Forecasting the next 3 months.\n'
    eai_preds = predict_EAI(new_components, 3)
    real_quarterly_data = real_quarterly_data.set_index('quarter').dropna(axis = 0, how = 'all').reset_index()

    log_str += 'Generating excel data.\n'
    filename = filepath +'/'+ model_name + '.xlsx'
    avg_cycles, eai_comps, eai_preds, eai_qtrly, gdp_hp_qtrly = prepare_country_file(eai_components, real_quarterly_data, cycle_stats, eai_preds, summary_dict, filename)
    
    log_str += 'Generating additional calculations and saving.\n'
    prepare_sheets(filename)


# ## Visualization

# In[60]:


### 1. EAI Components Bar Chart
def clean_df_bar(fileloc):
    df = pd.read_excel(fileloc, sheet_name=2)
    df['month'] = pd.to_datetime(df['month'])
    df = df.dropna(subset=['CONSUMPTION', 'EXO_INTERNATIONAL',
       'FINANCIAL', 'GOVERNMENT', 'INVESTMENT', 'TRADE'], how='all') 
    return df

def plotly_bar(df, x_bar, y_line, y_bar_list, plot_fig, write_html=False): 
    
    fig_bar = px.bar(
    df, x=x_bar, y=y_bar_list, title=f"{y_line} components"
    ).add_traces(
        px.line(df, x=x_bar, y=y_line).update_traces(showlegend=True, name=y_line).data
    ).update_layout(yaxis2={"side":"right", "overlaying":"y"}).update_layout(legend = dict(font = dict(family = "Courier", size = 15)),
                  legend_title = dict(font = dict(family = "Courier", size = 15)))
    if write_html == True:
        #save as html
        fig_bar.write_html(f"{y_line}_components_viz.html")
    if plot_fig == True:
        fig_bar.show()
    
    return fig_bar

### 2. Line Chart

def target_category(a, b):
    if (a == 0.3) and (b == 2.25):
        return 'increasing below trend, red'
    elif (a == -0.5) and (b == 1.35):
        return 'decreasing below trend, yellow'
    elif (a == 2.5) and (b == 1.35):
        return 'decreasing above trend, blue'
    elif (a == 1.7) and (b == 2.25):
        return 'increasing above trend, green'
    
def clean_df_line(fileloc):
    df_pred = pd.read_excel(fileloc, sheet_name=3)
    df_pred['month'] = pd.to_datetime(df_pred['month'])
    
    df_tag = pd.read_excel(fileloc, sheet_name=6)
    df_tag['target_tag'] = df_tag.apply(lambda x: target_category(x.x_param_eai, x.y_param_eai), 
                                        axis=1, result_type='expand')
    
    df_tag = df_tag[['month', 'target_tag']].dropna()
    df_tag['month'] = pd.to_datetime(df_tag['month'])
    df_tag[['target_tag', 'color']] = df_tag['target_tag'].str.split(',', expand=True)
    df_pred = df_pred.merge(df_tag, on='month', how='left')

    df_gdp = pd.read_excel(fileloc, sheet_name=5)
    df_gdp['qtr'] = pd.to_datetime(df_gdp['qtr'])
    df_pred = df_pred.merge(df_gdp[['qtr','GDP_Gap']], left_on='month', right_on='qtr', how='left').drop('qtr', axis=1)
    df_pred = df_pred[df_pred.month.dt.year>=2019].reset_index(drop=True)
        
    df_pred['legend'] = 'EAI'
    df_pred['legend2'] = 'Actual GDP Growth Gap'
    df_pred['month'] = pd.date_range(start=df_pred.month.iloc[2], freq='M', periods=len(df_pred.month)) #move to end of qtr

    return df_pred

def clean_df_line_2(df_pred):
    df_pred_ci_index = df_pred[(df_pred.EAI.isnull()) & ~(df_pred.PredictedEAI.isnull())].index.tolist()
    df_pred_ci_index.append(df_pred_ci_index[0]-1)
    df_pred_ci = df_pred.iloc[sorted(df_pred_ci_index)]

    df_pred_ci['legend3'] = '95%CiLowerLimit'
    df_pred_ci['legend4'] = '95%CiUpperLimit'
    df_pred_ci['legend5'] = '99%CiLowerLimit'
    df_pred_ci['legend6'] = '99%CiUpperLimit'
    df_pred_ci['legend7'] = 'PredictedEAI'
    
    return df_pred_ci

def plot_line(df_pred, df_pred_ci, fileloc, plot_fig): ##updated
#     ctry = df_pred['Country'].unique()[0]
    fig1 = px.line(df_pred, x="month", y="EAI", color='legend')
    fig2 = px.line(df_pred_ci, x="month", y="95%CiLowerLimit", color="legend3", color_discrete_sequence = ['pink'])
    fig3 = px.line(df_pred_ci, x="month", y="95%CiUpperLimit", color="legend4", color_discrete_sequence = ['pink'])
    fig4 = px.line(df_pred_ci, x="month", y="99%CiLowerLimit", color="legend5", color_discrete_sequence = ['brown'])
    fig5 = px.line(df_pred_ci, x="month", y="99%CiUpperLimit", color="legend6", color_discrete_sequence = ['brown'])
    fig6 = px.line(df_pred_ci, x="month", y="PredictedEAI", color="legend7", color_discrete_sequence = ['grey'])
    fig7 = px.scatter(df_pred, x="month", y="GDP_Gap", color='legend2', color_discrete_sequence=['black'])
    fig_line = go.Figure(data = fig1.data + fig2.data + fig3.data + fig4.data + fig5.data + fig6.data + fig7.data)
    time = df_pred.month.iloc[df_pred_ci.index[0]]
    fig_line.add_vline(x=time, line_color= 'grey', line_dash = 'dash')
    fig_line.add_annotation(x=time, y=1, yref="paper", text="Predicted Value")
    fig_line.update_layout(title = f"EAI and the Business Cycle").update_layout(legend = dict(font = dict(family = "Courier", size = 15)),
                  legend_title = dict(font = dict(family = "Courier", size = 15)))
    
    if plot_fig == True:
        fig_line.show()
    
    return fig_line

### 3. Dial Charts

def degree_range(n): 
    start = np.linspace(0,180,n+1, endpoint=True)[0:-1]
    end = np.linspace(0,180,n+1, endpoint=True)[1::]
    mid_points = start + ((end-start)/2.)
    return np.c_[start, end], mid_points

def rot_text(ang): 
    rotation = np.degrees(np.radians(ang) * np.pi / np.pi - np.radians(90))
    return rotation

def gauge(ax, labels=['LOW','MEDIUM','HIGH','VERY HIGH','EXTREME'], \
          colors='jet_r', arrow=1, title='', maintitle='', fname=False): 
    
    """
    some sanity checks first
    
    """
    
    N = len(labels)
    
    if arrow > N: 
        raise Exception("\n\nThe category ({}) is greated than \
        the length\nof the labels ({})".format(arrow, N))
 
    
    """
    if colors is a string, we assume it's a matplotlib colormap
    and we discretize in N discrete colors 
    """
    
    if isinstance(colors, str):
        cmap = cm.get_cmap(colors, N)
        cmap = cmap(np.arange(N))
        colors = cmap[::-1,:].tolist()
    if isinstance(colors, list): 
        if len(colors) == N:
            colors = colors[::-1]
        else: 
            raise Exception("\n\nnumber of colors {} not equal \
            to number of categories{}\n".format(len(colors), N))

    """
    begins the plotting
    """
    

    ang_range, mid_points = degree_range(N)

    labels = labels[::-1]
    
    """
    plots the sectors and the arcs
    """
    patches = []
    for ang, c in zip(ang_range, colors): 
        # sectors
        patches.append(Wedge((0.,0.), .4, *ang, facecolor='w', lw=2))
        # arcs
        patches.append(Wedge((0.,0.), .4, *ang, width=0.10, facecolor=c, lw=2, alpha=0.5))
    
    [ax.add_patch(p) for p in patches]

    
    """
    set the labels (e.g. 'LOW','MEDIUM',...)
    """

    for mid, lab in zip(mid_points, labels): 

        ax.text(0.35 * np.cos(np.radians(mid)), 0.35 * np.sin(np.radians(mid)), lab, \
            horizontalalignment='center', verticalalignment='center', fontsize=10, \
            fontweight='bold', rotation = rot_text(mid))

    """
    set the bottom banner and the title
    """
    r = Rectangle((-0.4,-0.1),0.8,0.1, facecolor='w', lw=2)
    ax.add_patch(r)
    ax.text(0, -0.08, title, horizontalalignment='center', \
         verticalalignment='center', fontsize=20,
           )
    
    """
    set the top banner and the title
    """
    r = Rectangle((-0.4,-0.1),0.8,0.1, facecolor='w', lw=2)
    ax.add_patch(r)
    ax.text(0, 0.45, maintitle, horizontalalignment='center', \
         verticalalignment='center', fontsize=20,
           )

    """
    plots the arrow now
    """
    if arrow==0:
        pass
    else:
        pos = mid_points[abs(arrow - N)]

        ax.arrow(0, 0, 0.225 * np.cos(np.radians(pos)), 0.225 * np.sin(np.radians(pos)), \
                     width=0.025, head_width=0.06, head_length=0.1, fc='k', ec='k')

        ax.add_patch(Circle((0, 0), radius=0.02, facecolor='k'))
        ax.add_patch(Circle((0, 0), radius=0.01, facecolor='w', zorder=11))

    """
    removes frame and ticks, and makes axis equal and tight
    """
    
    ax.set_frame_on(False)
    ax.axes.set_xticks([])
    ax.axes.set_yticks([])
    ax.axis('equal')
    plt.tight_layout()
    if fname:
        fig.savefig(fname, dpi=200)
        
def latest_dial_date(fileloc): 
    df = pd.read_excel(fileloc, sheet_name=9)
    df = df.dropna(subset=['EAI_quarterly']).reset_index(drop=True)
    df['quarter'] = pd.to_datetime(df['quarter'])
    quarter_number = df.iloc[-1]['quarter'].quarter

    df = pd.read_excel(fileloc, sheet_name=6)
    df['month'] = pd.to_datetime(df['month'])
    df = df.dropna(subset=['eai'])
    month = df.iloc[-1]['month'].month
    year = df.iloc[-1]['month'].year
    
    return quarter_number, month, year

def arrow_category(a, b):
    if (a == 0.3) and (b == 2.25):
        return 2
    elif (a == -0.5) and (b == 1.35):
        return 1
    elif (a == 2.5) and (b == 1.35):
        return 4
    elif (a == 1.7) and (b == 2.25):
        return 3
    else:
        return np.nan
    
def df_clean_quarterly_cycle(fileloc, quarter_number, year):
    df = pd.read_excel(fileloc, sheet_name=9)
    df = df.dropna(subset=['EAI_quarterly']).reset_index(drop=True)
#     ctry = df['Country'].unique()[0]
    df['quarter'] = pd.to_datetime(df['quarter'])
    
    if quarter_number == 1:
        quarter_number_month = '01'
    elif quarter_number == 2:
        quarter_number_month = '04'
    elif quarter_number == 3:
        quarter_number_month = '07'
    elif quarter_number == 4:
        quarter_number_month = '10'
    df_qtr = df[df.quarter == pd.to_datetime(f'{quarter_number_month}/01/{year}')]
    df_qtr['EAI_arrow_cat'] = df_qtr.apply(lambda x: arrow_category(x.x_param_eai_quarterly, x.y_param_eai_quarterly), 
                                axis=1, result_type='expand')

    eai_values = df_qtr[['prev_eai_quarterly', 'EAI_quarterly', 'EAI_arrow_cat']].values.flatten().tolist()
    eai_values.insert(0, 'EAI')
    df_qtr['gdp_gap_arrow_cat'] = df_qtr.apply(lambda x: arrow_category(x.x_param_gdp_gap_quarterly, x.y_param_gdp_gap_quarterly), 
                                axis=1, result_type='expand')
    gdp_gap_values = df_qtr[['prev_gdp_gap_quarterly', 'GDP_gap_quarterly', 'gdp_gap_arrow_cat']].values.flatten().tolist()
    gdp_gap_values.insert(0, 'GDP Growth Gap')
    df_qtr_cycle = pd.DataFrame([eai_values, gdp_gap_values], columns=['sub_title', 'prev', 'curr', 'arrow_cat'])

    return df_qtr_cycle 

def plot_quarterly_cycle(quarter_number, year, fileloc, df_qtr_cycle, plot_fig):

    df = pd.read_excel(fileloc, sheet_name=9)
    df = df.dropna(subset=['EAI_quarterly']).reset_index(drop=True)
#     ctry = df['Country'].unique()[0]

    fig, axes = plt.subplots(1, 2, figsize=(21.5, 10))
    axes = axes.flatten()
    for i, axe in enumerate(axes):
        df_row = df_qtr_cycle.iloc[i]
        maintitle, prev_val, curr_val, arrow_val = df_row[0], df_row[1], df_row[2], df_row[3]
        if np.isnan(prev_val):
            prev_val = 'None'
        else:
            prev_val = round(prev_val, 4)
        if np.isnan(curr_val):
            curr_val = 'None'
            arrow_val = 0
        else:
            curr_val = round(curr_val, 4)
            if np.isnan(arrow_val):
                arrow_val = 0
            else:
                arrow_val = int(arrow_val)
        
        gauge(ax=axe, labels=['BELOW TREND:\nDECREASING',
                  'BELOW TREND:\nINCREASING',
                  'ABOVE TREND:\nINCREASING',
                  'ABOVE TREND:\nDECREASING'], \
          colors=['#FFCC00','#ED1C24','green','cyan'], 
          arrow=arrow_val, 
          title=f'Prev: {prev_val}\nCurr: {curr_val}', maintitle=maintitle)
    fig.suptitle(f'Quarterly Business Cycle Data: Q{quarter_number} {year}', weight='bold', fontsize=20);
    if plot_fig == True:
        fig.show()
    else:
        plt.close(fig)
    return fig

def arrow_category_components(a):
    if a < 0:
        return 1 #below target
    elif a >= 0:
        return 2
    else:
        return np.nan
    
def df_sector_dials(fileloc, month, year):
    df = pd.read_excel(fileloc, sheet_name=9)
    df = df.dropna(subset=['EAI_quarterly']).reset_index(drop=True)
#     ctry = df['Country'].unique()[0]

    df = pd.read_excel(fileloc, sheet_name=6)
    df['month'] = pd.to_datetime(df['month'])
    df_month = df[df.month == pd.to_datetime(f'{month}/01/{year}')]
    df_month['EAI_arrow_cat'] = df_month.apply(lambda x: arrow_category(x.x_param_eai, x.y_param_eai), 
                                axis=1, result_type='expand')
    df_month.apply(lambda x: arrow_category_components(x.con), 
                                axis=1, result_type='expand')
    curr_cols = ['con', 'inv', 'tra', 'fin', 'gov', 'exo']
    for col in curr_cols:
        df_month[f'{col}_arrow'] = df_month.apply(lambda x: arrow_category_components(x[col]), 
                                axis=1, result_type='expand')
    eai_values = df_month[['eai', 'prev_eai', 'EAI_arrow_cat']].values.flatten().tolist()
    eai_values.insert(0, 'All Index')

    prev_cols = ['prev_con', 'prev_inv', 'prev_tra', 'prev_fin', 'prev_gov', 'prev_exo']
    sector_arrow = ['con_arrow', 'inv_arrow', 'tra_arrow', 'fin_arrow', 'gov_arrow', 'exo_arrow']
    sector_full = ['Consumption', 'Investment', 'Trade', 'Finance', 'Government', 'External Sector']
    df_curr_val = df_month[curr_cols].reset_index(drop=True).values.flatten().tolist()
    df_prev_val = df_month[prev_cols].reset_index(drop=True).values.flatten().tolist()
    df_sector_val = df_month[sector_arrow].reset_index(drop=True).values.flatten().tolist()
    return eai_values, sector_full, df_prev_val, df_curr_val, df_sector_val 

def plot_sector_dials(month, year, eai_values, sector_full, df_prev_val, df_curr_val, df_sector_val, plot_fig): 

    fig, axes = plt.subplots(2, 4, figsize=(21.5, 10))
    fig.delaxes(axes[1,3]) #delete last fig
    axes = axes.flatten()
    eai_name, curr_eai, prev_eai, eai_arrow = eai_values[0], eai_values[1], eai_values[2], eai_values[3]
    if np.isnan(prev_eai):
        prev_eai = 'None'
    else:
        prev_eai = round(prev_eai, 4)
    if np.isnan(curr_eai): #if nan, no arrow
        curr_eai = 'None'
        eai_arrow = 0
    else: #if not nan
        curr_eai = round(curr_eai, 4)
        if np.isnan(eai_arrow):
            eai_arrow = 0
        else:
            eai_arrow = int(eai_arrow)
    
    gauge(ax=axes[0], labels=['BELOW TREND:\nDECREASING',
                      'BELOW TREND:\nINCREASING',
                      'ABOVE TREND:\nINCREASING',
                      'ABOVE TREND:\nDECREASING'], \
              colors=['#FFCC00','#ED1C24','green','cyan'], 
              arrow=eai_arrow, 
              title=f'Prev: {prev_eai}\nCurr: {curr_eai}', maintitle=eai_name)
    
    for i, axe in enumerate(axes[1:-1]):
        sector_title = sector_full[i]
        prev_val = round(df_prev_val[i], 4)
        if np.isnan(prev_val):
            prev_val = 'None'
        curr_val = round(df_curr_val[i], 4)
        if np.isnan(curr_val):
                curr_val = 'None'
                arrow_val = 0
        else:        
            arrow_val = int(df_sector_val[i])
        gauge(ax=axe, labels=['BELOW TREND',
                  'ABOVE TREND'], \
          colors=['#7F00FF','#FF5F15'], 
          arrow=arrow_val, 
          title=f'Prev: {prev_val}\nCurr: {curr_val}', maintitle=sector_title)
    fig.suptitle(f'EAI and Sector / Category Dials: {calendar.month_name[month]} {year}', weight='bold', fontsize=15);
    
    if plot_fig == True:
        fig.show()
    else:
        plt.close(fig)
    return fig

### 4. Save to PDF

#plotly
def save_pdf_plotly(figures_list, filename, height=900): 
    
    image_list = [pio.to_image(fig, format='png', width=1440, height=height, scale=1.5) for fig in figures_list]
    for index, image in enumerate(image_list):
        with io.BytesIO() as tmp:
            tmp.write(image)  
            image = Image.open(tmp).convert('RGB')  # convert and overwrite 'image' to prevent creating a new variable
            image_list[index] = image  # overwrite byte image data in list, replace with PIL converted image data

    # pop first item from image_list, use that to access .save(). Then refer back to image_list to append the rest
    image_list.pop(0).save(filename, 'PDF',
                    save_all=True, append_images=image_list, resolution=100.0)  # TODO improve resolution
    
#plt
def save_pdf_plt(figures_list, filename):
    with PdfPages(filename) as pdf:
        for fig in figures_list:
            pdf.savefig(fig) 
            
def merge_pdf(pdf_list, filename):
    merger = PdfMerger()

    for pdf in pdf_list:
        merger.append(pdf)

    merger.write(filename)
    merger.close()
    
### 5. Merge All
def plot_EAI_dashboard(fileloc, plot_fig=False): 
    #only 2 bar charts and 1 line chart
    df = clean_df_bar(fileloc)
    fig_bar_EAI = plotly_bar(df, 'month', "EAI", ['CONSUMPTION', 'EXO_INTERNATIONAL',
               'FINANCIAL', 'GOVERNMENT', 'INVESTMENT', 'TRADE'], plot_fig)
    fig_bar_EAI_GDP = plotly_bar(df, 'month', 'EAI_GDP_gap', 
            ['CONSUMPTION_GDP_gap', 'EXO_INTERNATIONAL_GDP_gap',
           'FINANCIAL_GDP_gap', 'GOVERNMENT_GDP_gap', 'INVESTMENT_GDP_gap',
           'TRADE_GDP_gap'], plot_fig)
    df_pred = clean_df_line(fileloc)
    df_pred_ci = clean_df_line_2(df_pred)
    fig_line = plot_line(df_pred, df_pred_ci, fileloc, plot_fig)
    save_pdf_plotly([fig_bar_EAI, fig_bar_EAI_GDP, fig_line], f'{fileloc.replace("data.xlsx", "")}' + "EAI_dashboard.pdf")        

    
#2. Dial Charts
def plot_dial_charts_all(fileloc, plot_fig=False):
    quarter_number, month, year = latest_dial_date(fileloc)
    fig_list = []
    dates_list = pd.date_range(end=f'{month}/01/{year}', periods=36, freq='MS')
    quarter_prev = ''

    for date in dates_list:
        month = date.month
        quarter_number = date.quarter
        year = date.year
        df_qtr_cycle = df_clean_quarterly_cycle(fileloc, quarter_number, year) 
        eai_values, sector_full, df_prev_val, df_curr_val, df_sector_val = df_sector_dials(fileloc, month, year)
        if quarter_prev != quarter_number:
            fig_dial = plot_quarterly_cycle(quarter_number, year, fileloc, df_qtr_cycle, plot_fig)
            fig_dial_sector = plot_sector_dials(month, year, eai_values, sector_full, df_prev_val, df_curr_val, df_sector_val, plot_fig)
            fig_list.append(fig_dial)
            fig_list.append(fig_dial_sector)
            quarter_prev = quarter_number
        else:
            fig_dial_sector = plot_sector_dials(month, year, eai_values, sector_full, df_prev_val, df_curr_val, df_sector_val, plot_fig)
            fig_list.append(fig_dial_sector)
    save_pdf_plt(fig_list, f'{fileloc.replace("data.xlsx", "")}' + "EAI_dial_charts.pdf")
    
#3. Error
def clean_df(fileloc):
    df_eai = pd.read_excel(fileloc, sheet_name=4)
    df_eai['qtr'] = pd.to_datetime(df_eai['qtr'])

    df_gdp = pd.read_excel(fileloc, sheet_name=5)
    df_gdp['qtr'] = pd.to_datetime(df_gdp['qtr'])

    df = df_eai.merge(df_gdp, on='qtr')[['qtr', 'EAI', 'GDP_Gap']].dropna()
    return df

def compute_error(df, model_name):
    
    y_true = df.GDP_Gap.values.tolist()
    y_pred = df.EAI.values.tolist()

    mse = mean_squared_error(y_true, y_pred)
    rmse = mean_squared_error(y_true, y_pred, squared=False)
    mae = mean_absolute_error(y_true, y_pred)
    mape = mean_absolute_percentage_error(y_true, y_pred)
    r2 = r2_score(y_true, y_pred)

    df_res = pd.DataFrame([mse, rmse, mae, mape, r2], 
                          index=['MSE', 'RMSE', 'MAE', 'MAPE', 'R2'], 
                          columns=[model_name])
    return df_res

def df_error_metrics_ml(root, save_res=True):
    df_res_all = pd.DataFrame()
    model_files = model_filepath(root)
    for file in model_files:
        df = clean_df(file)
        df_res = compute_error(df, file) 
        df_res.columns = [file.replace("data.xlsx","").replace(f"{root}","").replace('\\', '').replace('/', '')]
        df_res_all = pd.concat([df_res_all, df_res], axis=1)
    df_res_all = df_res_all.T
    if save_res == True:
        df_res_all.to_csv(f"{root}/ml_error_metrics.csv")
    return df_res_all

def plot_error_ml(df_error, plot_fig=False):
    
    fig = make_subplots(rows=2, cols=2,
        subplot_titles=["RMSE", "MAE", "R2"],
        horizontal_spacing=0.37)

    fig.add_trace(go.Bar(x=df_error.RMSE, y=df_error.index, orientation='h'),
                  row=1, col=1)

    fig.add_trace(go.Bar(x=df_error.MAE, y=df_error.index, orientation='h'),
                  row=1, col=2)

    fig.add_trace(go.Bar(x=df_error.R2, y=df_error.index, orientation='h'),
                  row=2, col=1)

    fig.update_layout(height=500, width=1000,
                      title_text="Error Metrics Comparison", showlegend=False)

    fig.update_yaxes(type = 'category', categoryorder='max ascending') 
    
    if plot_fig == True:
        fig.show()
        
    return fig

def model_filepath(root):
    model_files = []
    for path, subdirs, files in os.walk(root):
        for name in files:
            if 'data.xlsx' in name:
                model_files.append(os.path.join(path, name))
    return model_files

def clean_df_pred_ml(root):
    model_files = model_filepath(root)
    df_all = clean_df(model_files[0])[['qtr', 'GDP_Gap', 'EAI']]
    df_all.rename(columns={"EAI":model_files[0].replace("data.xlsx","").replace(f"{root}","").replace('\\', '').replace('/', '')}, inplace=True) #set to model name
    for file in model_files[1:]:
        df = clean_df(file).drop('GDP_Gap', axis=1)
        df.rename(columns={"EAI":file.replace(f"{root}","").replace("data.xlsx","").replace('\\', '').replace('/', '')}, inplace=True) #set to model name
        df_all = df_all.merge(df, on='qtr', how='left')
    df_all.qtr = pd.date_range(start=df_all.qtr.iloc[0], freq='Q', periods=len(df_all.qtr)) 
    return df_all

def plot_line_predictions(df, plot_fig=False):
    fig_ml = px.line(df, x='qtr', y=df.columns[2:])

    fig_gdpgap = go.Figure([go.Scatter(
            name='GDP Gap',
            x=df['qtr'],
            y=df['GDP_Gap'],
            line=dict(color="black", width=3.5),
            mode='lines+markers',
            showlegend=True),
            ])
    fig3 = go.Figure(data=fig_gdpgap.data + fig_ml.data)
    fig3.update_layout(
        title='EAI Predictions and GDP Gap',
    )
    
    if plot_fig == True:
        fig3.show()
        
    return fig3

def plot_EAI_pred_error(output_path, plot_fig=False):
    #line graph
    df = clean_df_pred_ml(output_path)
    fig_pred = plot_line_predictions(df, plot_fig)
    
    #error
    df_error = df_error_metrics_ml(output_path, save_res=True)  
    fig_error = plot_error_ml(df_error, plot_fig)
    
    #save all
    save_pdf_plotly([fig_pred, fig_error], f"{output_path}/EAI_predictions_comparison.pdf", 1100)


# ### Additional UI Screens

# In[61]:


#Browse folders for filepath to save all output files
#Status update screen (run once process flow is finished)
#Status update screen should run all processing, modeling, and final saving functions


# In[62]:


#open the progress screen and run all back end functions
def UI_progress_screen(event, model_selection, added_data_dict, date_dict, monthly_info, quarterly_info, output_path):
    
    cancel_button = sg.Button('Cancel', key='CANCEL_OPERATIONS', enable_events = True, font=('Calibri', 8), expand_x=False)
    title_txt = sg.Text('Running operations...', font=('Calibri Bold', 14), expand_x=True, justification='center')
    instructions_txt = sg.Text('Please wait while EAI predictions are being calculated. If any errors occur, an error message will be displayed. Verify your data inputs carefully. If the issue continues, contact the development team.', font=('Calibri', 10), expand_x=True, justification='center')
    status_txt = sg.Multiline(key='TEXTBOX', write_only=True, size=(60,10), font=('Calibri', 8), reroute_cprint=True, expand_x=True, autoscroll=True, auto_refresh=True,)
    
    model_ls = []
    for key in model_selection.keys():
        if model_selection[key] == True:
            model_ls.append(key)
    
    progress_ele = sg.ProgressBar(len(model_ls), key='PROGRESS_BAR', size=(60,5), bar_color='white', expand_x=True)
    
    layout = [
        [title_txt],
        [instructions_txt],
        [status_txt],
        [progress_ele],
        [cancel_button]
        ]
    
    window = sg.Window('Economic Activity Index: Running Operations', layout, grab_anywhere=True, resizable=True)
    event = 'START'
    filepath = output_path
    
    for i in range(len(model_ls)):
        event, values = window.read(timeout=0)
        if event == 'CANCEL_OPERATIONS' or event == None:
            break
        else:
            exception_msg = ''
            try:
                model_name = model_ls[i]
                model = full_model_dictionary[model_name]

                sg.cprint('Preparing data: ' + model_name +' model.', font=('Calibri Bold', 8))
                full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict = prep_data(added_data_dict, monthly_info, quarterly_info)
                exception_msg = 'Data preparation failed.'


                sg.cprint('Generating predictions.', font='Calibri 8')
                full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly = generate_ml_predictions(full_df, old_data_monthly, old_data_quarterly, cycles_dict, contributions_dict, date_dict, model)
                exception_msg = 'ML prediction generation failed.'

                sg.cprint('Calculating initial components.', font='Calibri 8')
                eai_comps, rmse_df, ovr_weights, calcs_quarterly, cycle_stats = get_eai_components(full_df, calcs_monthly, calcs_quarterly, stat_dict, data_monthly)
                exception_msg = 'Component calculation failed.'

                sg.cprint('Calculating data summaries.', font='Calibri 8')
                new_rmse_df, summary_dict = calculate_quarterly_summaries(rmse_df)
                exception_msg = 'Data summary calculation failed.'

                sg.cprint('Calculating overall components.', font='Calibri 8')
                eai_components = calculate_overall_components(eai_comps, summary_dict, ovr_weights)
                real_quarterly_data = calcs_quarterly[['quarter','HPBasedCycle','CFBasedCycle','HamiltonBasedCycle','GDP_Gap']]

                try:
                    eai_components['date'] = pd.to_datetime(eai_components['date'])
                    eai_components = eai_components.sort_values(by=['date'], ascending = True)
                except:
                    pass
                new_components = eai_components.copy()
                comp_ls = list(new_components['EAI'])
                count = 0
                for i in range(len(new_components)):
                    if math.isnan(comp_ls[-i]):
                        count+=1
                    else:
                        break
                new_components = new_components[:-(count-1)]
                exception_msg = 'Overall component calculations failed.'

                sg.cprint('Forecasting the next 3 months.', font='Calibri 8')
                eai_preds = predict_EAI(new_components, 3)
                real_quarterly_data = real_quarterly_data.set_index('quarter').dropna(axis = 0, how = 'all').reset_index()
                exception_msg = 'Forecasting failed.'

                sg.cprint('Generating excel data.', font='Calibri 8')
                
                
                Path(filepath+'/'+model_name).mkdir(parents=True, exist_ok=True)
                
                filename = filepath +'/'+ model_name + '/'+'data.xlsx'
                avg_cycles, eai_comps, eai_preds, eai_qtrly, gdp_hp_qtrly = prepare_country_file(eai_components, real_quarterly_data, cycle_stats, eai_preds, summary_dict, filename)
                exception_msg = 'Generating excel data failed..'

                sg.cprint('Generating additional calculations and saving.', font='Calibri 8')
                prepare_sheets(filename)
                exception_msg = 'Additional calculations and file saving failed.'
                
                #viz per model
                sg.cprint('Generating visualizations and saving.', font='Calibri 8')
                plot_EAI_dashboard(filename, date_dict)
                plot_dial_charts_all(filename)
                exception_msg = 'Creating visualizations failed.'
            except Exception as e:
                error_message = repr(e)
                #sg.cprint(exception_msg +': '+error_message, font='Calibri 8')
                sg.cprint('Operations Failed: '+error_message, font='Calibri 8')
            window['PROGRESS_BAR'].update_bar(i+1, len(model_ls))
    try:
        sg.cprint('Generating comparative model analysis.', font=('Calibri Bold', 8))
        plot_EAI_pred_error(output_path, date_dict)
    except:
        sg.cprint('Comparative model analysis generation failed.', font=('Calibri Bold', 8))
    window.close()


# In[66]:


#run initial GUI process to gather all relevant information before running back end functions
event, model_selection, added_data_dict, date_dict, monthly_info, quarterly_info, output_path = UI_process_flow()


# In[64]:


#run back end functions and show progress screen
UI_progress_screen(event, model_selection, added_data_dict, date_dict, monthly_info, quarterly_info, output_path)


# In[ ]:




